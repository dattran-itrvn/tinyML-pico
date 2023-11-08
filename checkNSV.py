
import csv
import numpy as np
import wfdb
from matplotlib import pyplot as plt
from preprocess import preprocess_data
from define import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import wfdb
import scipy.signal as signal
from scipy.signal import butter, filtfilt, lfilter, iirnotch


def butter_lowpass(cutoff, fs, order=3):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = butter(order, normal_cutoff, btype='low', analog=False)
    return b, a


def butter_lowpass_filter(data, cutoff, fs, order=3):
    b, a = butter_lowpass(cutoff, fs, order=order)
    y = filtfilt(b, a, data)
    return y


def butter_highpass(cutoff, fs, order=3):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = butter(order, normal_cutoff, btype='high', analog=False)
    return b, a


def butter_highpass_filter(data, cutoff, fs, order=3):
    b, a = butter_highpass(cutoff, fs, order=order)
    y = filtfilt(b, a, data)
    return y


def butter_bandpass(lowcut, highcut, fs, order=3):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    b, a = butter(order, [low, high], btype='band')
    return b, a


def butter_bandpass_filter(data, lowcut, highcut, fs, order=3):
    b, a = butter_bandpass(lowcut, highcut, fs, order=order)
    y = filtfilt(b, a, data)
    return y


def find_max_locate(ecg_array):
    arr = ecg_array
    y_i = np.amax(arr)
    i_1, = np.where(arr == y_i)
    x_i = i_1[0]
    return y_i, i_1, x_i


def derivative_filter(fs):
    if fs != 200:
        int_c = (4 - 0) / (fs * (1 / 40))
        x = np.arange(0, 4)
        xp = np.array([1, 2, 0, -2, -1]) * (1 / 8) * fs
        fp = np.linspace(0, int_c, 5)  # mang 5 gia tri buoc nhay int_c
        b = np.interp(x, xp, fp)  # 1-D data interpolation (trả về mảng 1 chiều nội suy)b = np.interp(x, xp, fp)
    else:
        b = np.array([1, 2, 0, -2, -1]) * (1 / 8) * fs
    return b


# def pan_tompkin(ecg, fs):
#     qrs_c = []#np.zeros(shape=(1,1)) # amplitude of R
#     qrs_i = []  # index
#     #SIG_LEV = 0
#     nois_c = []
#     nois_i = []
#     # delay = 0
#     # skip = 0
#     # not_nois = 0
#     # selected_RR = []
#     m_selected_RR = 0
#     mean_RR = 0
#     qrs_i_raw = []
#     qrs_amp_raw = []
#     ser_back = []
#     ecg = np.nan_to_num(ecg)
#     if (fs == 200):
#         ecg = ecg - np.mean(ecg)  # remove mean of Signal
#         #Low pass filter
#         ecg_l = butter_lowpass_filter(data=ecg, cutoff_low = 12, fs = fs, order=3)
#         ecg_l = ecg_l / (np.max(abs(ecg_l)))
#         #Highpass filter
#         ecg_h = butter_highpass_filter(data = ecg_l, cutoff_high = 5, fs = fs, order=3)
#         ecg_h = ecg_h / np.max(abs(ecg_h))
#     else:
#         #Bandpass filter for Noise cancelation of other sampling frequencies(Filtering)
#         ecg_h = butter_bandpass_filter(data = ecg, lowcut = 5, highcut = 15, fs = fs, order=3)
#         ecg_h = ecg_h / np.max(abs(ecg_h))
#
#     #Derivative filter H(z) = (1/8T)(-z^(-2) - 2z^(-1) + 2z + z^(2))#######################
#     b = derivative_filter(fs)
#     ecg_d = signal.filtfilt(b, 1, ecg_h)
#     ecg_d = ecg_d / np.max(ecg_d)
#
#     # Squaring nonlinearly enhance the dominant peaks
#     ecg_s = ecg_d ** 2
#
#     #Moving average Y(nt) = (1/N)[x(nT-(N - 1)T)+ x(nT - (N - 2)T)+...+x(nT)]
#     window_len = round(0.150 * fs)
#     if window_len % 2 == 0:
#           window_len += 1
#     s = np.r_[ecg_s[window_len - 1:0:-1], ecg_s, ecg_s[-2:-window_len-1:-1]]# add element into a array. ex.between ecg_s
#     ecg_m = np.convolve(s, np.ones(window_len), mode='valid') / round(0.150 * fs)
#     ecg_m= ecg_m[int(window_len / 2):-int(window_len / 2)]
#
#     #Find all peaks in ecg signal
#     # ecg_m[ecg_m < 0.0001] = 0
#     # locs = peakutils.indexes(ecg_m, thres=0.5, min_dist=fs*0.2)
#     locs = signal.find_peaks(ecg_m, threshold=0.00000000001, distance=fs*0.22)[0]
#     pks = ecg_m[locs]
#     # Initialize the training phase (2 seconds of the signal) to determine the THR_SIG and THR_NOISE
#     THR_SIG = (np.max(ecg_m[0:2 * fs])) * 1 / 3  # 0.25 of the max amplitude
#     THR_NOISE = (np.mean(ecg_m[0:2 * fs])) * 1 / 2  # 0.5 of the mean signal is considered to be noise
#     SIG_LEV = THR_SIG
#     NOISE_LEV = THR_NOISE
#     # Initialize bandpass filter threshold(2 seconds of the bandpass signal)
#     THR_SIG1 = (np.max(ecg_h[0:2 * fs])) * 1 / 3
#     THR_NOISE1 = (np.mean(ecg_h[0:2 * fs])) * 1 / 2
#     SIG_LEV1 = THR_SIG1
#     NOISE_LEV1 = THR_NOISE1
#
#
#     for i in range(0, len(pks)):
#         # locate the corresponding peak in the filtered signal
#         if (locs[i] - int(round(0.150 * fs))) >= 1 and locs[i] <= len(ecg_h)-1:
#             y_i, i_l, x_i = find_max_locate(ecg_h[(locs[i] - int(round(0.150 * fs))):(locs[i])])
#         else:
#             if i == 0:
#                 y_i, i_1, x_i = find_max_locate(ecg_h[0:locs[i]])
#                 ser_back = 1
#             elif locs[i] >= len(ecg_h):
#                 y_i, i_2, x_i = find_max_locate(ecg_h[(locs[i] - round(0.150 * fs)):-1])
#
#     # update the number of beats (Two heart rate means one the most recent and the other selector)
#         x = len(qrs_c)
#         # calculate the mean of the last 8 R waves to make sure that QRS is not
#         # missing(If no R detected, trigger a search back) 1.66*mean
#         if  x >= 9:
#             x1 = np.asarray(qrs_i[x-9: x])
#             diffRR = np.diff(x1)  # calculate RR interval
#             mean_RR = np.mean(diffRR)  # calculater the mean of 8 previous R waves interval
#             comp = qrs_i[-1] - qrs_i[-2]  # lastest RR
#             if comp <= 0.92 * mean_RR or comp >= 1.16 * mean_RR:  # most recent RR interval that fell between the accpetable low and high RR-interval limits
#                 # lower down thresholds to detect better in MVI
#                 THR_SIG = 0.5 * (THR_SIG)
#                 # lower down thresholds to detect better in Bandpass filtered
#                 THR_SIG1 = 0.5 * (THR_SIG1)
#             else:
#                 m_selected_RR = mean_RR  # the lastest regular beats mean
#
#         if m_selected_RR:
#             test_m = m_selected_RR  # if the regular RR available use it
#         elif mean_RR and m_selected_RR == 0:
#             test_m = mean_RR
#         else:
#             test_m = 0
#
#         if test_m:
#             if ((locs[i]-qrs_i[-1]) >= round(1.66 * test_m)): # it shows a QRS is missed
#                 try:
#                     pks_temp, i_3, locs_temp = find_max_locate(ecg_m[(qrs_i[-1] + round(0.2 * fs)):(locs[i] - round(0.2 * fs))])
#                 except Exception as e:
#                     print("{} get error: {}".format(i, e))
#                 locs_temp = qrs_i[-1] + round(0.2 * fs) + locs_temp - 1  # location
#                 if pks_temp > THR_NOISE:
#                     qrs_c.append(pks_temp)
#                     qrs_i.append(locs_temp)
#                 # find the location in filtered sig
#                 if locs_temp <= len(ecg_h):
#                     y_i_t, i_4, x_i_t = find_max_locate(ecg_h[(locs_temp - round(0.150 * fs)):locs_temp])
#                 else:
#                     y_i_t, i_5, x_i_t = find_max_locate(ecg_h[(locs_temp - round(0.150 * fs)):locs_temp])
#                 # take care of bandpass signal threshold
#                 if y_i_t > THR_NOISE1:
#                     qrs_i_raw.append(locs_temp - round(0.150 * fs) + (x_i_t - 1))  # save index of bandpass
#                     qrs_amp_raw.append(y_i_t)  # save amplitude of bandpass
#                     SIG_LEV1 = 0.25 * y_i_t + 0.75 * SIG_LEV1  # when found with the second thres
#
#                 not_nois = 1
#                 SIG_LEV = 0.25 * pks_temp + 0.75 * SIG_LEV
#             else:
#                 not_nois = 0
#
#         # find noise and QRS peaks
#         if pks[i] >= THR_SIG:
#             # if a QRS candidate occurs within 360ms of the previous QRS
#             # the algorithm determines if its T wave or QRS
#             skip = 0
#             if len(qrs_c) >= 3:
#                 if (locs[i] - qrs_i[-1]) <= round(0.3600 * fs):
#                     Slope1 = np.mean(np.diff(ecg_m[locs[i] - round(0.075 * fs):locs[i]])) # mean slope of the waveform at that position
#                     Slope2 = np.mean(np.diff(ecg_m[(qrs_i[-1] - round(0.075 * fs)): qrs_i[-1]]))  # mean slope of previous R wave
#                     if abs(Slope1) <= abs(0.5 * Slope2):  # slope less then 0.5 of previous R
#                         nois_c.append(pks[i])
#                         nois_i.append(locs[i])
#                         skip = 1  # T wave identification
#                         # adjust noise level in both filtered and MVI
#                         NOISE_LEV1 = 0.125 * y_i + 0.875 * NOISE_LEV1
#                         NOISE_LEV = 0.125 * pks[i] + 0.875 * NOISE_LEV
#                     # else:
#                     #     skip = 0
#             if skip == 0:  # skip is 1 when a T wave is detected
#                 qrs_c.append(pks[i])
#                 qrs_i.append(locs[i])
#                 # bandpass filter check threshold
#                 if y_i >= THR_SIG1:
#                     if ser_back:
#                         qrs_i_raw.append(x_i)
#                     else:
#                         qrs_i_raw.append(locs[i] - round(0.150 * fs) + (x_i - 1))
#                 qrs_amp_raw.append(y_i)# save amplitude of bandpass
#                 SIG_LEV1 = 0.125 * y_i + 0.875 * SIG_LEV1 # adjust threshold for bandpass filtered sig
#
#             # adjust Signal level
#             SIG_LEV = 0.125 * pks[i] + 0.875 * SIG_LEV
#
#         elif (THR_NOISE <= pks[i]) and (pks[i] < THR_SIG):
#             # adjust Noise level in filtered sig
#             NOISE_LEV1 = 0.125 * y_i + 0.875 * NOISE_LEV1
#             # adjust Noise level in MVI
#             NOISE_LEV = 0.125 * pks[i] + 0.875 * NOISE_LEV
#
#         elif pks[i] < THR_NOISE:
#             nois_c.append(pks[i])
#             nois_i.append(locs[i])
#             NOISE_LEV1 = 0.125 * y_i + 0.875 * NOISE_LEV1
#             NOISE_LEV = 0.125 * pks[i] + 0.875 * NOISE_LEV
#
#         # adjust the threshold with SNR
#         if NOISE_LEV != 0 or SIG_LEV != 0:
#             THR_SIG = NOISE_LEV + 0.25 * (abs(SIG_LEV - NOISE_LEV))
#             THR_NOISE = 0.5 * (THR_SIG)
#
#         # adjust the threshold with SNR for bandpassed signal
#         if NOISE_LEV1 != 0 or SIG_LEV1 != 0:
#             THR_SIG1 = NOISE_LEV1 + 0.25 * (abs(SIG_LEV1 - THR_NOISE1))
#             THR_NOISE1 = 0.5 * (THR_SIG1)
#
#     skip = 0  # reset parameters
#     not_nois = 0  # reset parameters
#     ser_back = 0  # reset bandpass param
#
#     symbol = []
#     for i in range(0, len(qrs_i)):
#         symbol.append('N')
#     # locs1 = []
#     # for e in locs:
#     #     locs1.append(e)
#
#     qrs_i1 = []
#     for e in qrs_i:
#         qrs_i1.append(e - 0)
#
#     return qrs_i1, symbol

# def main(DIR, namefile):
#     record = wfdb.rdrecord(DIR + str(namefile))
#     ecg_o = record.p_signal[:, 1]
#     d_ptk_bts, symbol_ptk =  pan_tompkin(ecg_o,fs= record.fs)
#     ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
#     bts = np.zeros_like(beats)
#     # btss = np.zeros_like(d_ptk_bts)
#     # for i in range(len(d_ptk_bts)):
#     #     btss[i] = np.argmax(
#     #         ecg_o[d_ptk_bts[i] - int(FREQUENCY_SAMPLING * BXB_WIN):d_ptk_bts[i] + int(FREQUENCY_SAMPLING * BXB_WIN)]) + d_ptk_bts[
#     #                  i] - int(FREQUENCY_SAMPLING * BXB_WIN)
#     for i in range(len(beats)):
#         bts[i] = np.argmax(
#             ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)]) + beats[
#                      i] - int(FREQUENCY_SAMPLING * BXB_WIN)
#     sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
#     sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'S'
#     sym[(sym == 'V') | (sym == 'e')] = 'V'
#     # sym[(sym == 'f') | (sym == 'Q')] = 'O'
#     sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'
#     symidx = np.zeros(len(sym))
#     nondata = sym.copy()
#     nondata[:] = ' '
#     symidx[sym != 'N'] = 1
#     fill_colorV = np.where(sym == 'V')[0] + 2
#     fill_colorS = np.where(sym == 'S')[0] + 2
#     detectsym = sym.copy()
#     detectsym[1:-1] = ''
#     detectsym2 = sym.copy()
#     detectsym2[2:-2] = ''
#
#     df = np.diff(bts)
#     idx = 1
#     divba = np.zeros(len(beats))
#     divba[1:-1] = df[:-1] / df[1:]
#     while (idx < len(sym) - 1):
#         cnt = 0
#         while (divba[idx] > divba[idx + 1] and idx < len(sym) - 2):
#             idx += 1
#             cnt += 1
#         if cnt == 0:
#             detectsym[idx] = 'N'
#         elif divba[idx] < 0.6:
#             for k in range(cnt):
#                 if divba[idx - cnt + k] >= 1.27:
#                     detectsym[idx - cnt + k] = 'N'
#                 else:
#                     detectsym[idx - cnt + k] = 'V'
#             detectsym[idx] = 'V'
#         else:
#             detectsym[idx - cnt:idx + 1] = 'N'
#         idx += 1
#
#     for i in range(2,len(sym)-2):
#         if divba[i-1]>=1.2:
#             if divba[i]<0.7:
#                 if divba[i-1]>=2:
#                     detectsym2[i]='V'
#                 elif divba[i+1]>=1:
#                     detectsym2[i]='V'
#                 else:
#                     detectsym2[i]='N'
#             else:
#                 detectsym2[i]='N'
#         else:
#             detectsym2[i]='N'
#
#
#
#     sym2 = sym.copy()
#     # sym2[np.where(sym2=='O'|sym2=='F')[0]]='N'
#     idx2 = 0
#     while(idx2<len(sym)):
#         cnt=0
#         while(sym2[idx2]!='N'):
#             cnt+=1
#             idx2+=1
#             if idx2 >= len(sym):
#                 break
#         if cnt==1:
#             sym2[idx2-1]='V'
#         else:
#             sym2[idx2-cnt:idx2]='N'
#         idx2 += 1
#
#
#     nN = len(np.where(sym[1:-1] == 'N')[0])
#     nSV = len(sym[1:-1]) - nN
#
#     nSV6 = len(np.nonzero(divba[1:-1] < 0.6)[0])
#     nTSV6 = len(np.nonzero(sym[np.nonzero(divba[1:-1] < 0.6)[0] + 1] == 'S')[0]) + len(
#         np.nonzero(sym[np.nonzero(divba[1:-1] < 0.6)[0] + 1] == 'V')[0])
#     nFSV6 = nSV6 - nTSV6
#     nN6 = len(np.nonzero(divba[1:-1] >= 0.6)[0])
#     nTN6 = len(np.nonzero(sym[np.nonzero(divba[1:-1] >= 0.6)[0] + 1] == 'N')[0]) + len(
#         np.nonzero(sym[np.nonzero(divba[1:-1] >= 0.6)[0] + 1] == 'O')[0]) + len(
#         np.nonzero(sym[np.nonzero(divba[1:-1] >= 0.6)[0] + 1] == 'F')[0])
#     nFN6 = nN6 - nTN6
#
#     detectsym6 = sym.copy()
#     detectsym6[np.where(divba[1:-1] < 0.6)[0]+1]='V'
#     detectsym6[np.where(divba[1:-1] >= 0.6)[0]+1]='N'
#
#     detectsym85 = sym.copy()
#     detectsym85[np.where(divba[1:-1] < 0.85)[0] + 1] = 'V'
#     detectsym85[np.where(divba[1:-1] >= 0.85)[0] + 1] = 'N'
#
#     symfullV=sym.copy()
#     symfullV[np.where(sym == "S")[0]]='V'
#
#
#     # N6=np.where(divba[1:-1] < 0.6)[0]+1
#
#
#     if nTSV6 + nFN6 == 0:
#         SESV6 = 0
#     else:
#         SESV6 = round(nTSV6 / (nTSV6 + nFN6), 5)
#     if nTSV6 + nFSV6 == 0:
#         PSV6 = 0
#     else:
#         PSV6 = round(nTSV6 / (nTSV6 + nFSV6), 5)
#     if nTN6 + nFSV6 == 0:
#         SEN6 = 0
#     else:
#         SEN6 = round(nTN6 / (nTN6 + nFSV6), 5)
#     if nTN6 + nFN6 == 0:
#         PN6 = 0
#     else:
#         PN6 = round(nTN6 / (nTN6 + nFN6), 5)
#
#     # nSV85=len(np.nonzero(divba[1:-1]<0.85)[0])
#     # nN85=len(np.nonzero(divba[1:-1]>=0.85)[0])
#
#     nSV85 = len(np.nonzero(divba[1:-1] < 0.85)[0])
#     nTSV85 = len(np.nonzero(sym[np.nonzero(divba[1:-1] < 0.85)[0] + 1] == 'S')[0]) + len(
#         np.nonzero(sym[np.nonzero(divba[1:-1] < 0.85)[0] + 1] == 'V')[0])
#     nFSV85 = nSV85 - nTSV85
#     nN85 = len(np.nonzero(divba[1:-1] >= 0.85)[0])
#     nTN85 = len(np.nonzero(sym[np.nonzero(divba[1:-1] >= 0.85)[0] + 1] == 'N')[0]) + len(
#         np.nonzero(sym[np.nonzero(divba[1:-1] >= 0.85)[0] + 1] == 'O')[0]) + len(
#         np.nonzero(sym[np.nonzero(divba[1:-1] >= 0.85)[0] + 1] == 'F')[0])
#     nFN85 = nN85 - nTN85
#
#     if nTSV85 + nFN85 == 0:
#         SESV85 = 0
#     else:
#         SESV85 = round(nTSV85 / (nTSV85 + nFN85), 5)
#     if nTSV85 + nFSV85 == 0:
#         PSV85 = 0
#     else:
#         PSV85 = round(nTSV85 / (nTSV85 + nFSV85), 5)
#     if nTN85 + nFSV85 == 0:
#         SEN85 = 0
#     else:
#         SEN85 = round(nTN85 / (nTN85 + nFSV85), 5)
#     if nTN85 + nFN85 == 0:
#         PN85 = 0
#     else:
#         PN85 = round(nTN85 / (nTN85 + nFN85), 5)
#
#     cl = [SEN6 < 0.8, PN6 < 0.8, SESV6 < 0.8, PSV6 < 0.8, SEN85 < 0.8, PN85 < 0.8, SESV85 < 0.8, PSV85 < 0.8]
#     ThD = [namefile, ' ', len(sym[1:-1]), ' ', nN, nSV, ' ', nN6, nTN6, nFN6, SEN6, PN6, ' ', nSV6, nTSV6, nFSV6, SESV6,
#            PSV6, ' ', nN85, nTN85, nFN85, SEN85, PN85, ' ', nSV85, nTSV85, nFSV85, SESV85, PSV85]
#
#     err = []
#     err.extend(np.where(detectsym == 'V')[0][np.where(
#         (sym[np.where(detectsym == 'V')[0]] == 'N') | (sym[np.where(detectsym == 'V')[0]] == 'O') | (
#                     sym[np.where(detectsym == 'V')[0]] == 'F'))[0]] + 2)
#     err.extend(np.where((detectsym == 'N') | (detectsym == 'O') | (detectsym == 'F'))[0][np.where(
#         (sym[np.where((detectsym == 'N') | (detectsym == 'O') | (detectsym == 'F'))[0]] == 'S') | (
#         (sym[np.where((detectsym == 'N') | (detectsym == 'O') | (detectsym == 'F'))[0]] == 'V')))[0]] + 2)
#     err2 = []
#     err2.extend(np.where(detectsym2 == 'V')[0][np.where(
#         (sym2[np.where(detectsym2 == 'V')[0]] == 'N') | (sym2[np.where(detectsym2 == 'V')[0]] == 'O') | (
#                 sym2[np.where(detectsym2 == 'V')[0]] == 'F'))[0]] + 2)
#     err2.extend(np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0][np.where(
#         (sym2[np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0]] == 'S') | (
#             (sym2[np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0]] == 'V')))[0]] + 2)
#
#     # expt = np.where((divba >= 0.6) & (divba < 0.85))[0] + 2
#     expt = np.where((divba >= 1.2))[0] + 2
#     expt2 = np.where((divba >= 1))[0] + 2
#
#     minlocal1 = []
#     minlocal2 = []
#     QRv = []
#     QRs = []
#     RSv = []
#     RSs = []
#     QSv = []
#     QSs = []
#     QS2 = []
#     QS5 = []
#     QS8 = []
#     print(namefile)
#     for i in range(len(bts)):
#         # print(i)
#         if bts[i] < 40:
#             minlocal1.append(0 + np.argmin(ecg[0:bts[i]]))
#         else:
#             minlocal1.append(bts[i] - 40 + np.argmin(ecg[bts[i] - 40:bts[i]]))
#
#         if bts[i] > len(ecg) - 40:
#             minlocal2.append(bts[i] + np.argmin(ecg[bts[i]:]))
#         else:
#             minlocal2.append(bts[i] + np.argmin(ecg[bts[i]:bts[i] + 40]))
#
#         QSv.append(ecg[minlocal2[i]] - ecg[minlocal1[i]])
#         QSs.append(minlocal2[i] - minlocal1[i])
#
#         QRv.append(ecg[bts[i]] - ecg[minlocal1[i]])
#         QRs.append(bts[i] - minlocal1[i])
#
#         RSv.append(ecg[bts[i]] - ecg[minlocal2[i]])
#         RSs.append(minlocal2[i] - bts[i])
#
#         sRS = 0.2 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
#         sQR = 0.2 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
#         QS2.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
#                     np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
#
#         sRS = 0.5 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
#         sQR = 0.5 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
#         QS5.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
#                     np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
#
#         sRS = 0.8 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
#         sQR = 0.8 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
#         QS8.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
#                     np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
#         # mor_diff_f = []
#         # mor_diff_f.append(0)
#         # for i in range(1,len(QSv)-2):
#         #     mor_diff_f.append()
#     s=0
#     # return sym, bts, np.round(beats / FREQUENCY_SAMPLING,
#     #                           3),divba, QSv, QSs, QRv, QRs, RSv, RSs, QS2, QS5, QS8, fill_colorV, fill_colorS, nondata,err, expt,expt2, err2
#     # annotation = wfdb.Annotation(record_name=str(namefile),
#     #                              extension='atr',
#     #                              sample=np.asarray(bts),
#     #                              symbol=np.asarray(sym2),
#     #                              fs=360,
#     #                              )
#     # wfdb.wrann(record_name=str(namefile),
#     #                              extension='atr',
#     #                              sample=np.asarray(bts),
#     #                              symbol=np.asarray(symfullV),
#     #                              fs=360,
#     #                  write_dir='ann2')
#
#     # wfdb.wrann(record_name=str(namefile),
#     #            extension='atr',
#     #            sample=np.asarray(bts),
#     #            symbol=np.asarray(sym2),
#     #            fs=360,
#     #            write_dir='detect_SV_single')
#     #
#     # wfdb.wrann(record_name=str(namefile),
#     #            extension='txt',
#     #            sample=np.asarray(bts),
#     #            symbol=np.asarray(detectsym2),
#     #            fs=360,
#     #            write_dir='detect_SV_single')
#
#
#     # wfdb.wrann(record_name=str(namefile),
#     #            extension='txt',
#     #            sample=np.asarray(bts),
#     #            symbol=np.asarray(detectsym6),
#     #            fs=360,
#     #            write_dir='annNSV6')
#     #
#     wfdb.wrann(record_name=str(namefile),
#                extension='atr',
#                sample=np.asarray(beats),
#                symbol=np.asarray(sym),
#                fs=360,
#                write_dir='PTK')
#
#     wfdb.wrann(record_name=str(namefile),
#                extension='txt',
#                sample=np.asarray(d_ptk_bts),
#                symbol=np.asarray(symbol_ptk),
#                fs=360,
#                write_dir='PTK')

def pan_tompkin(ecg, fs):
    qrs_c = []  # np.zeros(shape=(1,1)) # amplitude of R
    qrs_i = []  # index
    # SIG_LEV = 0
    nois_c = []
    nois_i = []
    # delay = 0
    # skip = 0
    # not_nois = 0
    # selected_RR = []
    m_selected_RR = 0
    mean_RR = 0
    qrs_i_raw = []
    qrs_amp_raw = []
    ser_back = []
    ecg = np.nan_to_num(ecg)
    if (fs == 200):
        ecg = ecg - np.mean(ecg)  # remove mean of Signal
        # Low pass filter
        ecg_l = butter_lowpass_filter(data=ecg, cutoff_low=12, fs=fs, order=3)
        ecg_l = ecg_l / (np.max(abs(ecg_l)))
        # Highpass filter
        ecg_h = butter_highpass_filter(data=ecg_l, cutoff_high=5, fs=fs, order=3)
        ecg_h = ecg_h / np.max(abs(ecg_h))
    else:
        # Bandpass filter for Noise cancelation of other sampling frequencies(Filtering)
        ecg_h = butter_bandpass_filter(data=ecg, lowcut=5, highcut=15, fs=fs, order=3)
        ecg_h = ecg_h / np.max(abs(ecg_h))

    # Derivative filter H(z) = (1/8T)(-z^(-2) - 2z^(-1) + 2z + z^(2))#######################
    b = derivative_filter(fs)
    ecg_d = signal.filtfilt(b, 1, ecg_h)
    ecg_d = ecg_d / np.max(ecg_d)

    # Squaring nonlinearly enhance the dominant peaks
    ecg_s = ecg_d ** 2
    # plt.plot(ecg_s)
    # Moving average Y(nt) = (1/N)[x(nT-(N - 1)T)+ x(nT - (N - 2)T)+...+x(nT)]
    window_len = round(0.150 * fs)
    if window_len % 2 == 0:
        window_len += 1
    s = np.r_[
        ecg_s[window_len - 1:0:-1], ecg_s, ecg_s[-2:-window_len - 1:-1]]  # add element into a array. ex.between ecg_s
    ecg_m = np.convolve(s, np.ones(window_len), mode='valid') / round(0.150 * fs)
    ecg_m = ecg_m[int(window_len / 2):-int(window_len / 2)]

    # Find all peaks in ecg signal
    # ecg_m[ecg_m < 0.0001] = 0
    # locs = peakutils.indexes(ecg_m, thres=0.5, min_dist=fs*0.2)
    locs = signal.find_peaks(ecg_m, threshold=0.00000000001, distance=fs * 0.22)[0]
    pks = ecg_m[locs]
    # plt.plot(ecg_m)
    # plt.show()
    # Initialize the training phase (2 seconds of the signal) to determine the THR_SIG and THR_NOISE
    THR_SIG = (np.max(ecg_m[0:2 * fs])) * 1 / 3  # 0.25 of the max amplitude
    THR_NOISE = (np.mean(ecg_m[0:2 * fs])) * 1 / 2  # 0.5 of the mean signal is considered to be noise
    SIG_LEV = THR_SIG
    NOISE_LEV = THR_NOISE
    # Initialize bandpass filter threshold(2 seconds of the bandpass signal)
    THR_SIG1 = (np.max(ecg_h[0:2 * fs])) * 1 / 3
    THR_NOISE1 = (np.mean(ecg_h[0:2 * fs])) * 1 / 2
    SIG_LEV1 = THR_SIG1
    NOISE_LEV1 = THR_NOISE1

    for i in range(0, len(pks)):
        # locate the corresponding peak in the filtered signal
        if (locs[i] - int(round(0.150 * fs))) >= 1 and locs[i] <= len(ecg_h) - 1:
            y_i, i_l, x_i = find_max_locate(ecg_h[(locs[i] - int(round(0.150 * fs))):(locs[i])])
        else:
            if i == 0:
                y_i, i_1, x_i = find_max_locate(ecg_h[0:locs[i]])
                ser_back = 1
            elif locs[i] >= len(ecg_h):
                y_i, i_2, x_i = find_max_locate(ecg_h[(locs[i] - round(0.150 * fs)):-1])

        # update the number of beats (Two heart rate means one the most recent and the other selector)
        x = len(qrs_c)
        # calculate the mean of the last 8 R waves to make sure that QRS is not
        # missing(If no R detected, trigger a search back) 1.66*mean
        if x >= 9:
            x1 = np.asarray(qrs_i[x - 9: x])
            diffRR = np.diff(x1)  # calculate RR interval
            mean_RR = np.mean(diffRR)  # calculater the mean of 8 previous R waves interval
            comp = qrs_i[-1] - qrs_i[-2]  # lastest RR
            if comp <= 0.92 * mean_RR or comp >= 1.16 * mean_RR:  # most recent RR interval that fell between the accpetable low and high RR-interval limits
                # lower down thresholds to detect better in MVI
                THR_SIG = 0.5 * (THR_SIG)
                # lower down thresholds to detect better in Bandpass filtered
                THR_SIG1 = 0.5 * (THR_SIG1)
            else:
                m_selected_RR = mean_RR  # the lastest regular beats mean

        if m_selected_RR:
            test_m = m_selected_RR  # if the regular RR available use it
        elif mean_RR and m_selected_RR == 0:
            test_m = mean_RR
        else:
            test_m = 0

        if test_m:
            if ((locs[i] - qrs_i[-1]) >= round(1.66 * test_m)):  # it shows a QRS is missed
                try:
                    pks_temp, i_3, locs_temp = find_max_locate(
                        ecg_m[(qrs_i[-1] + round(0.2 * fs)):(locs[i] - round(0.2 * fs))])
                except Exception as e:
                    print("{} get error: {}".format(i, e))
                locs_temp = qrs_i[-1] + round(0.2 * fs) + locs_temp - 1  # location
                if pks_temp > THR_NOISE:
                    qrs_c.append(pks_temp)
                    qrs_i.append(locs_temp)
                # find the location in filtered sig
                if locs_temp <= len(ecg_h):
                    y_i_t, i_4, x_i_t = find_max_locate(ecg_h[(locs_temp - round(0.150 * fs)):locs_temp])
                else:
                    y_i_t, i_5, x_i_t = find_max_locate(ecg_h[(locs_temp - round(0.150 * fs)):locs_temp])
                # take care of bandpass signal threshold
                if y_i_t > THR_NOISE1:
                    qrs_i_raw.append(locs_temp - round(0.150 * fs) + (x_i_t - 1))  # save index of bandpass
                    qrs_amp_raw.append(y_i_t)  # save amplitude of bandpass
                    SIG_LEV1 = 0.25 * y_i_t + 0.75 * SIG_LEV1  # when found with the second thres

                not_nois = 1
                SIG_LEV = 0.25 * pks_temp + 0.75 * SIG_LEV
            else:
                not_nois = 0

        # find noise and QRS peaks
        if pks[i] >= THR_SIG:
            # if a QRS candidate occurs within 360ms of the previous QRS
            # the algorithm determines if its T wave or QRS
            skip = 0
            if len(qrs_c) >= 3:
                if (locs[i] - qrs_i[-1]) <= round(0.3600 * fs):
                    Slope1 = np.mean(np.diff(
                        ecg_m[locs[i] - round(0.075 * fs):locs[i]]))  # mean slope of the waveform at that position
                    Slope2 = np.mean(
                        np.diff(ecg_m[(qrs_i[-1] - round(0.075 * fs)): qrs_i[-1]]))  # mean slope of previous R wave
                    if abs(Slope1) <= abs(0.5 * Slope2):  # slope less then 0.5 of previous R
                        nois_c.append(pks[i])
                        nois_i.append(locs[i])
                        skip = 1  # T wave identification
                        # adjust noise level in both filtered and MVI
                        NOISE_LEV1 = 0.125 * y_i + 0.875 * NOISE_LEV1
                        NOISE_LEV = 0.125 * pks[i] + 0.875 * NOISE_LEV
                    # else:
                    #     skip = 0
            if skip == 0:  # skip is 1 when a T wave is detected
                qrs_c.append(pks[i])
                qrs_i.append(locs[i])
                # bandpass filter check threshold
                if y_i >= THR_SIG1:
                    if ser_back:
                        qrs_i_raw.append(x_i)
                    else:
                        qrs_i_raw.append(locs[i] - round(0.150 * fs) + (x_i - 1))
                qrs_amp_raw.append(y_i)  # save amplitude of bandpass
                SIG_LEV1 = 0.125 * y_i + 0.875 * SIG_LEV1  # adjust threshold for bandpass filtered sig

            # adjust Signal level
            SIG_LEV = 0.125 * pks[i] + 0.875 * SIG_LEV

        elif (THR_NOISE <= pks[i]) and (pks[i] < THR_SIG):
            # adjust Noise level in filtered sig
            NOISE_LEV1 = 0.125 * y_i + 0.875 * NOISE_LEV1
            # adjust Noise level in MVI
            NOISE_LEV = 0.125 * pks[i] + 0.875 * NOISE_LEV

        elif pks[i] < THR_NOISE:
            nois_c.append(pks[i])
            nois_i.append(locs[i])
            NOISE_LEV1 = 0.125 * y_i + 0.875 * NOISE_LEV1
            NOISE_LEV = 0.125 * pks[i] + 0.875 * NOISE_LEV

        # adjust the threshold with SNR
        if NOISE_LEV != 0 or SIG_LEV != 0:
            THR_SIG = NOISE_LEV + 0.25 * (abs(SIG_LEV - NOISE_LEV))
            THR_NOISE = 0.5 * (THR_SIG)

        # adjust the threshold with SNR for bandpassed signal
        if NOISE_LEV1 != 0 or SIG_LEV1 != 0:
            THR_SIG1 = NOISE_LEV1 + 0.25 * (abs(SIG_LEV1 - THR_NOISE1))
            THR_NOISE1 = 0.5 * (THR_SIG1)

    skip = 0  # reset parameters
    not_nois = 0  # reset parameters
    ser_back = 0  # reset bandpass param

    symbol = []
    for i in range(0, len(qrs_i)):
        symbol.append('N')
    # locs1 = []
    # for e in locs:
    #     locs1.append(e)

    qrs_i1 = []
    for e in qrs_i:
        qrs_i1.append(e - 0)
    # print('length beats_detection: ', len(qrs_i1))
    # print('beats_detection:', qrs_i1)
    return qrs_i1, symbol


def main(DIR, namefile):
    record = wfdb.rdrecord(DIR + str(namefile))
    ecg_o = record.p_signal[:, 0]
    d_ptk_bts, symbol_ptk = pan_tompkin(ecg_o, fs=record.fs)

    ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
    bts = np.zeros_like(beats)
    print(namefile)
    for i in range(len(beats)):
        if ecg[beats[i]] < 0:
            # print(i, ' ', 1)
            bts[i] = np.argmax(
                -(ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
        else:
            bts[i] = np.argmax(
                (ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
            # print(i, ' ', 2)
    l_btss = len(d_ptk_bts)
    top = 0
    bot = 0
    if d_ptk_bts[0] - 30 < 0:
        l_btss -= 1
        symbol_ptk = symbol_ptk[1:]
        top = 1
    if d_ptk_bts[-1] + 30 > len(ecg):
        l_btss -= 1
        symbol_ptk = symbol_ptk[:-1]
        bot = 1
    btss = np.zeros(l_btss, dtype='int64')

    for j in range(len(d_ptk_bts) - bot - top):
        btss[j] = np.argmax(abs(ecg[d_ptk_bts[j + top] - 30:d_ptk_bts[j + top] + 30])) + d_ptk_bts[j + top] - 30



    fill_colorS = []
    symidx = np.zeros(len(sym))

    symidx[sym != 'N'] = 1



    minlocal1 = []
    minlocal2 = []
    QS2 = []
    QS5 = []
    QS8 = []
    diff_f2 = [0]
    diff_a2 = []
    diff_f5 = [0]
    diff_a5 = []
    diff_f8 = [0]
    diff_a8 = []
    diff_fm = [0]
    diff_am = []
    diff_tb = []
    print(namefile)
    for i in range(len(btss)):
        # print(i)
        if ecg[btss[i]] >= 0:
            d = 1
        else:
            d = -1

        if btss[i] < 30:
            s1= 0 + np.argmin(d * ecg[0:btss[i]])
        else:
            s1 = btss[i] - 30 + np.argmin(d * ecg[btss[i] - 30:btss[i]])

        if btss[i] > len(ecg) - 30:
            s2 = btss[i] + np.argmin(d * ecg[btss[i]:])
        else:
           s2 = btss[i] + np.argmin(d * ecg[btss[i]:btss[i] + 30])
        if ecg[s1]*ecg[s2]>=0:
            if abs(ecg[s1]) > abs(ecg[s2]):
                s12 = s1
            else:
                s12 = s2
        else:
            if ecg[s1]*ecg[btss[i]]<0:
                s12 = s1
            else:
                s12 = s2

        diff_tb.append(round(abs(ecg[btss][i]-ecg[s12]),6))

        cnt = 0
        f = True
        while (ecg[btss[i] - cnt] * ecg[btss[i]] > 0):
            cnt += 1
            if btss[i] - cnt < 0 or cnt > 30:
                minlocal1.append(btss[i] - cnt + 1)
                f = False
                break
        if f:
            minlocal1.append(btss[i] - cnt)
        cnt = 0
        f = True
        while (ecg[btss[i] + cnt] * ecg[btss[i]] > 0):
            cnt += 1
            if btss[i] + cnt > len(ecg)-1 or cnt > 30:
                minlocal2.append(btss[i] + cnt - 1)
                f = False
                break
        if f:
            minlocal2.append(btss[i] + cnt)
        # print(i)
        sRS = 0.8 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.8 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS2.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.5 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.5 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS5.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.2 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.2 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS8.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
        if i>0 and i < len(btss):
            diff_f2.append(round(abs((QS2[i]-QS2[i-1])/QS2[i-1]),6))
            diff_a2.append(round(abs((QS2[i-1]-QS2[i])/QS2[i]),6))
            diff_f5.append(round(abs((QS5[i] - QS5[i - 1])/QS5[i - 1]),6))
            diff_a5.append(round(abs((QS5[i - 1] - QS5[i])/QS5[i]),6))
            diff_f8.append(round(abs((QS8[i] - QS8[i - 1])/QS8[i - 1]),6))
            diff_a8.append(round(abs((QS8[i - 1] - QS8[i])/QS8[i]),6))
            diff_fm.append(round((diff_f2[i]+diff_f5[i]+diff_f8[i])/3,6))
            diff_am.append(round((diff_a2[i-1]+diff_a5[i-1]+diff_a8[i-1])/3,6))
        # if i==len(btss)-1:
    diff_a2.append(0)
    diff_a5.append(0)
    diff_a8.append(0)
    diff_am.append(0)

    QS2=np.array(QS2)
    QS5=np.array(QS5)
    QS8=np.array(QS8)
    diff_tb=np.array(diff_tb)
    diff_a2=np.array(diff_a2)
    diff_a5=np.array(diff_a5)
    diff_a8=np.array(diff_a8)
    diff_f2 = np.array(diff_f2)
    diff_f5 = np.array(diff_f5)
    diff_f8 = np.array(diff_f8)
    err = []

    df = np.diff(np.array(btss))
    divba = np.zeros(len(btss))
    divba[1:-1] = df[:-1] / df[1:]
    expt= []
    # expt.extend(np.where(np.array(diff_fm)<0.15)[0] + 2)
    expt2 = []
    # expt2.extend(np.where(np.array(diff_am) < 0.15)[0] + 2)
    err2 = []
    # err2.extend(np.where(divba>1.4)[0]+2)
    rr = [0]
    rr.extend(df)
    d_S = []
    d_V = []
    symbol_ptk_np = np.array(symbol_ptk)
    idx = 0
    while (idx < len(btss) - 1):
        if rr[idx + 1] <= 100 and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.8 and abs(
                ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.8:
            btss = np.delete(btss, idx)
            df = np.diff(np.array(btss))
            rr = [0]
            rr.extend(df)
            rr = np.array(rr)
            symbol_ptk_np = np.delete(symbol_ptk_np, idx)
            QS2 = np.delete(QS2, idx)
            QS5 = np.delete(QS5, idx)
            QS8 = np.delete(QS8, idx)
            diff_tb = np.delete(diff_tb, idx)
            diff_a2 = np.delete(diff_a2, idx)
            diff_a5 = np.delete(diff_a5, idx)
            diff_a8 = np.delete(diff_a8, idx)
            diff_f2 = np.delete(diff_f2, idx)
            diff_f5 = np.delete(diff_f5, idx)
            diff_f8 = np.delete(diff_f8, idx)
            continue
        if idx != 0 and rr[idx] <= 100 and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.8 and abs(
                ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.8:
            btss = np.delete(btss, idx)
            df = np.diff(np.array(btss))
            rr = [0]
            rr.extend(df)
            rr = np.array(rr)
            symbol_ptk_np = np.delete(symbol_ptk_np, idx)
            QS2 = np.delete(QS2, idx)
            QS5 = np.delete(QS5, idx)
            QS8 = np.delete(QS8, idx)
            diff_tb = np.delete(diff_tb, idx)
            diff_a2 = np.delete(diff_a2, idx)
            diff_a5 = np.delete(diff_a5, idx)
            diff_a8 = np.delete(diff_a8, idx)
            diff_f2 = np.delete(diff_f2, idx)
            diff_f5 = np.delete(diff_f5, idx)
            diff_f8 = np.delete(diff_f8, idx)
            continue
        # if rr[idx] <= 100 and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.8 and abs(ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.8:
        #     btss = np.delete(btss,idx)
        #     rr = np.delete(rr,idx)
        #     symbol_ptk_np = np.delete(symbol_ptk_np,idx)
        #     continue
        idx += 1
    idx = 1
    while (idx < len(btss) - 1):
        if rr[idx] <= 100:
            btss = np.delete(btss, idx - 1)
            df = np.diff(np.array(btss))
            rr = [0]
            rr.extend(df)
            rr = np.array(rr)
            symbol_ptk_np = np.delete(symbol_ptk_np, idx - 1)
            continue
        idx+=1
    nondata = symbol_ptk_np.copy()
    nondata[:] = ' '
    time = btss / FREQUENCY_SAMPLING
    fill_colorV = []
    s = []
    s.extend(np.where(np.array(rr)<=100)[0]+2)
    print(s)
    return sym, btss, np.round(time,6), np.round(divba,6),rr,np.round(ecg[btss],6),diff_tb, QS2, QS5, QS8, diff_f2, diff_a2,diff_f5, diff_a5, diff_f8, diff_a8, diff_fm, diff_am, fill_colorV, fill_colorS, nondata, err, expt, expt2, err2, d_V, d_S

def checkPTK(DIR, namefile):
        record = wfdb.rdrecord(DIR + str(namefile))
        ecg_o = record.p_signal[:, 0]
        d_ptk_bts, symbol_ptk = pan_tompkin(ecg_o, fs=record.fs)

        ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
        bts = np.zeros_like(beats)
        print(namefile)
        for i in range(len(beats)):
            if ecg[beats[i]] < 0:
                # print(i, ' ', 1)
                bts[i] = np.argmax(
                    -(ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                         beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
            else:
                bts[i] = np.argmax(
                    (ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                         beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
                # print(i, ' ', 2)
        l_btss = len(d_ptk_bts)
        top = 0
        bot = 0
        if d_ptk_bts[0] - 30 < 0:
            l_btss -= 1
            symbol_ptk = symbol_ptk[1:]
            top = 1
        if d_ptk_bts[-1] + 30 > len(ecg):
            l_btss -= 1
            symbol_ptk = symbol_ptk[:-1]
            bot = 1
        btss = np.zeros(l_btss, dtype='int64')

        for j in range(len(d_ptk_bts) - bot - top):
            btss[j] = np.argmax(abs(ecg[d_ptk_bts[j + top] - 30:d_ptk_bts[j + top] + 30])) + d_ptk_bts[j + top] - 30

        sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
        sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'N'
        sym[(sym == 'V') | (sym == 'e')] = 'N'
        sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'

        symidx = np.zeros(len(sym))
        nondata = sym.copy()
        nondata[:] = ' '
        symidx[sym != 'N'] = 1

        detectsym2 = np.array(symbol_ptk.copy())
        detectsym2[:2] = 'N'
        detectsym2[-2:] = 'N'

        df = np.diff(np.array(btss))
        if df[0] < 200:
            df[0] = 200
        rr = [0]
        rr.extend(df)
        rr = np.array(rr)
        divba = np.zeros(len(btss))
        divba[1:-1] = df[:-1] / df[1:]


        # turn 2 detect SV

        minlocal1 = []
        minlocal2 = []
        QS2 = []
        QS5 = []
        QS8 = []
        diff_f2 = [0]
        diff_a2 = []
        diff_f5 = [0]
        diff_a5 = []
        diff_f8 = [0]
        diff_a8 = []
        diff_fm = [0]
        diff_am = []
        diff_tb = []
        # print(namefile)
        for i in range(len(btss)):
            # print(i)
            if ecg[btss[i]] >= 0:
                d = 1
            else:
                d = -1

            if btss[i] < 30:
                s1 = 0 + np.argmin(d * ecg[0:btss[i]])
            else:
                s1 = btss[i] - 30 + np.argmin(d * ecg[btss[i] - 30:btss[i]])

            if btss[i] > len(ecg) - 30:
                s2 = btss[i] + np.argmin(d * ecg[btss[i]:])
            else:
                s2 = btss[i] + np.argmin(d * ecg[btss[i]:btss[i] + 30])
            if ecg[s1] * ecg[s2] >= 0:
                if abs(ecg[s1]) > abs(ecg[s2]):
                    s12 = s1
                else:
                    s12 = s2
            else:
                if ecg[s1] * ecg[btss[i]] < 0:
                    s12 = s1
                else:
                    s12 = s2

            diff_tb.append(round(abs(ecg[btss][i] - ecg[s12]), 6))

            cnt = 0
            f = True
            while (ecg[btss[i] - cnt] * ecg[btss[i]] > 0):
                cnt += 1
                if btss[i] - cnt < 0 or cnt > 30:
                    minlocal1.append(btss[i] - cnt + 1)
                    f = False
                    break
            if f:
                minlocal1.append(btss[i] - cnt)
            cnt = 0
            f = True
            while (ecg[btss[i] + cnt] * ecg[btss[i]] > 0):
                cnt += 1
                if btss[i] + cnt > len(ecg) - 1 or cnt > 30:
                    minlocal2.append(btss[i] + cnt - 1)
                    f = False
                    break
            if f:
                minlocal2.append(btss[i] + cnt)
            # print(i)
            sRS = 0.8 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
            sQR = 0.8 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
            QS2.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                    np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

            sRS = 0.5 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
            sQR = 0.5 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
            QS5.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                    np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

            sRS = 0.2 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
            sQR = 0.2 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
            QS8.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                    np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
            if i > 0 and i < len(btss) - 1:
                diff_f2.append(round(abs((QS2[i] - QS2[i - 1]) / QS2[i - 1]), 6))
                diff_a2.append(round(abs((QS2[i - 1] - QS2[i]) / QS2[i]), 6))
                diff_f5.append(round(abs((QS5[i] - QS5[i - 1]) / QS5[i - 1]), 6))
                diff_a5.append(round(abs((QS5[i - 1] - QS5[i]) / QS5[i]), 6))
                diff_f8.append(round(abs((QS8[i] - QS8[i - 1]) / QS8[i - 1]), 6))
                diff_a8.append(round(abs((QS8[i - 1] - QS8[i]) / QS8[i]), 6))
                diff_fm.append(round((diff_f2[i] + diff_f5[i] + diff_f8[i]) / 3, 6))
                diff_am.append(round((diff_a2[i - 1] + diff_a5[i - 1] + diff_a8[i - 1]) / 3, 6))
            if i == len(btss) - 1:
                diff_a2.append(0)
                diff_a5.append(0)
                diff_a8.append(0)
                diff_am.append(0)


        # classify S and V
        QS2 = np.array(QS2)
        QS5 = np.array(QS5)
        QS8 = np.array(QS8)
        symbol_ptk_np = np.array(symbol_ptk)
        idx=0
        while (idx<len(btss)-1):
            if rr[idx+1] <= int(FREQUENCY_SAMPLING*0.2) and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.8 and abs(ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.8:
                btss = np.delete(btss,idx)
                df = np.diff(np.array(btss))
                rr = [0]
                rr.extend(df)
                rr = np.array(rr)
                symbol_ptk_np = np.delete(symbol_ptk_np,idx)
                continue
            if idx!=0 and rr[idx] <= int(FREQUENCY_SAMPLING*0.2) and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.8 and abs(ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.8:
                btss = np.delete(btss,idx)
                df = np.diff(np.array(btss))
                rr = [0]
                rr.extend(df)
                rr = np.array(rr)
                symbol_ptk_np = np.delete(symbol_ptk_np,idx)
                continue
            # if rr[idx] <= 100 and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.8 and abs(ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.8:
            #     btss = np.delete(btss,idx)
            #     rr = np.delete(rr,idx)
            #     symbol_ptk_np = np.delete(symbol_ptk_np,idx)
            #     continue
            idx+=1
        # for i in range(1,len(btss)-1):
        idx=1
        while (idx<len(btss)-1):
            if rr[idx] <= int(FREQUENCY_SAMPLING*0.2):
                btss = np.delete(btss,idx-1)
                df = np.diff(np.array(btss))
                rr = [0]
                rr.extend(df)
                rr = np.array(rr)
                symbol_ptk_np = np.delete(symbol_ptk_np,idx-1)
                continue

            idx+=1
        # idx = 1
        # while (idx < len(btss) - 1):
        #     if rr[idx] <= 150 and abs(ecg[btss[idx]] / ecg[btss[idx - 1]]) < 0.7 and abs(ecg[btss[idx]] / ecg[btss[idx + 1]]) < 0.7:
        #         btss = np.delete(btss, idx)
        #         df = np.diff(np.array(btss))
        #         rr = [0]
        #         rr.extend(df)
        #         rr = np.array(rr)
        #         symbol_ptk_np = np.delete(symbol_ptk_np, idx)
        #         continue
        #
        #     idx += 1

        wfdb.wrann(record_name=str(namefile),
                   extension='atr',
                   sample=np.asarray(bts),
                   symbol=np.asarray(sym),
                   fs=360,
                   write_dir='PTK')

        wfdb.wrann(record_name=str(namefile),
                   extension='txt',
                   sample=np.asarray(btss),
                   symbol=np.asarray(symbol_ptk_np),
                   fs=360,
                   write_dir='PTK')
def ann(DIR, namefile):
    record = wfdb.rdrecord(DIR + str(namefile))
    ecg_o = record.p_signal[:, 0]
    d_ptk_bts, symbol_ptk = pan_tompkin(ecg_o, fs=record.fs)

    ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
    bts = np.zeros_like(beats)
    print(namefile)
    for i in range(len(beats)):
        if ecg[beats[i]] < 0:
            # print(i, ' ', 1)
            bts[i] = np.argmax(
                -(ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
        else:
            bts[i] = np.argmax(
                (ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
            # print(i, ' ', 2)
    l_btss = len(d_ptk_bts)
    top = 0
    bot = 0
    if d_ptk_bts[0] - 30 < 0:
        l_btss -= 1
        symbol_ptk = symbol_ptk[1:]
        top = 1
    if d_ptk_bts[-1] + 30 > len(ecg):
        l_btss -= 1
        symbol_ptk = symbol_ptk[:-1]
        bot = 1
    btss = np.zeros(l_btss, dtype='int64')

    for j in range(len(d_ptk_bts) - bot - top):
        btss[j] = np.argmax(abs(ecg[d_ptk_bts[j + top] - 30:d_ptk_bts[j + top] + 30])) + d_ptk_bts[j + top] - 30

    # sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
    # sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'S'
    # sym[(sym == 'V') | (sym == 'e')] = 'V'
    # sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'
    sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
    sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'S'
    sym[(sym == 'V') | (sym == 'e')] = 'V'
    sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'
    fill_colorV = np.where(sym == 'V')[0] + 2
    fill_colorS = np.where(sym == 'S')[0] + 2
    symidx = np.zeros(len(sym))
    nondata = sym.copy()
    nondata[:] = ' '
    symidx[sym != 'N'] = 1

    # detectsym2 = symbol_ptk.copy()
    #
    # # detectsym2[2:len(detectsym2)-1] = ' '
    #
    # df = np.diff(np.array(btss))
    # divba = np.zeros(len(btss))
    # divba[1:-1] = df[:-1] / df[1:]
    #
    # for i in range(2, len(btss) - 2):
    #     if divba[i - 1] >= 1.2:
    #         if divba[i] < 0.7:
    #             if divba[i - 1] >= 2:
    #                 detectsym2[i] = 'V'
    #             elif divba[i + 1] >= 1:
    #                 detectsym2[i] = 'V'
    #             else:
    #                 detectsym2[i] = 'N'
    #         else:
    #             detectsym2[i] = 'N'
    #     else:
    #         detectsym2[i] = 'N'

    detectsym2 = sym.copy()
    df = np.diff(np.array(bts))
    divba = np.zeros(len(bts))
    divba[1:-1] = df[:-1] / df[1:]

    rr = [0]
    rr.extend(df)
    rr = np.array(rr)
    sym2 = sym.copy()

    idx2 = 0
    while (idx2 < len(sym)):
        cnt = 0
        while (sym2[idx2] != 'N'):
            cnt += 1
            idx2 += 1
            if idx2 >= len(sym):
                break
        if cnt == 1:
            # sym2[idx2 - 1] = 'V'
            sym2[idx2 - 1] = sym[idx2-1]
        else:
            sym2[idx2 - cnt:idx2] = 'N'
        idx2 += 1


    flag_N = False
    flag_cnt=False
    if namefile==119 or namefile==207 or namefile==208:
        idx = 2  # 119
    elif namefile==232:
        idx = 3
    else:
        idx = 1

    cnt= 0
    x=0
    # detectsym2[0]='N'
    # detectsym2[1:-1]=''
    while (divba[idx]>0.6 and (abs(divba[idx+1]-divba[idx])<0.3) and divba[idx]<1.4):
        idx+=1
        cnt+=1
    # if np.mean(df[1:idx+1])>300:
    #     detectsym2[1:idx+1]='N'
    # if np.mean(rr[idx-cnt:idx+1])>300:
    #     detectsym2[idx-cnt:idx+1]='N'
    detectsym2[idx - cnt:idx + 1] = 'N'
    idx+=1
    cnt+=1
    s=0
    s_tmp=-1
    # if rr[idx + 1] / np.mean(rr[idx-cnt:idx + 1]) >= 0.8 and rr[idx+1]/np.mean(rr[idx-cnt:idx+1]) <= 1.2:
    #     detectsym2[idx + 1] = 'N'
    #     if rr[idx+2]/np.mean(rr[idx-cnt:idx+2]) >= 0.8 and rr[idx+2]/np.mean(rr[idx-cnt:idx+2]) <= 1.2:
    #         detectsym2[idx+2]='N'
    #     else:
    #         detectsym2[idx + 2] = 'D'
    # elif rr[idx + 1] / np.mean(rr[idx-cnt:idx + 1]) > 1.2:
    #     detectsym2[idx + 1] = 'N'
    #     if rr[idx + 2] / np.mean(rr[idx-cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[idx-cnt:idx + 1]) <= 1.2:
    #         detectsym2[idx + 1:idx + 3] = 'N'
    #     else:
    #         detectsym2[idx + 2] = 'D'
    while(True):
        if flag_N:

            cnt=0
            # tmp_while = np.mean(rr[idx - cnt:idx])
            while (divba[idx] > 0.6 and (abs(divba[idx+1]-divba[idx]) < 0.3) and divba[idx] < 1.4):
                # print(idx)
                idx += 1
                cnt += 1
                if idx >= len(divba) - 1:
                    break
            # if cnt==1:
            #     cnt=0
            # else:
            detectsym2[idx - cnt:idx + 1] = 'N'

            # idx+=1
            flag_N=False
        else:
            cnt_tmp = cnt
            if detectsym2[idx - 1] == 'N' and cnt_tmp>0:
                if idx==42:
                    m=0
                tmp = np.mean(rr[idx - cnt_tmp:idx])
                if rr[idx]/tmp >= 0.7:
                    detectsym2[idx] = 'N'
                    if rr[idx]/tmp>3:
                        rr[idx]=tmp
                    idx += 1
                    cnt += 1
                    tmp = np.mean(rr[idx - cnt:idx])
                    if idx >= len(divba) - 1:
                        break
                    # flag_N =True
                    # cnt_tmp = cnt
                    if rr[idx]/tmp >= 0.7:
                        detectsym2[idx] = 'N'
                        if rr[idx] / tmp > 3:
                            rr[idx] = tmp
                        if divba[idx] < 1.4:
                            flag_N = True
                            continue
                        # if rr[idx] / tmp >= 0.65:
                        #     detectsym2[idx] = 'N'
                        #     flag_N = True
                        # else:
                        #     detectsym2[idx] = 'D'
                        #     idx -= 1
                        #     cnt -= 1
                            # cnt_tmp=cnt+1
                    else:
                        if rr[idx] / tmp >= 0.6 and s_tmp > 0 and rr[idx] / s_tmp >= 1.3:
                            detectsym2[idx] = 'N'
                            if divba[idx] < 1.4:
                                flag_N = True
                                continue
                        else:
                            if rr[idx]>=300:
                                detectsym2[idx] = 'N'
                            else:
                                detectsym2[idx] = 'V'
                    # else den noi de so sanh RR với mean{RR} truoc
                else:
                    if rr[idx]/tmp >= 0.6 and s_tmp>0 and rr[idx]/s_tmp >= 1.3:
                        detectsym2[idx] = 'N'
                        # if rr[idx] / tmp < 1.3:
                        #     tmp = np.mean(rr[idx - cnt_tmp:idx])
                        idx += 1
                        cnt += 1
                        tmp = np.mean(rr[idx - cnt:idx])
                        if idx >= len(divba) - 1:
                            break
                        # cnt_tmp = cnt

                        if rr[idx] / tmp >= 0.7:
                            detectsym2[idx] = 'N'
                            if rr[idx] / tmp > 3:
                                rr[idx] = tmp
                            if divba[idx] < 1.4:
                                flag_N = True
                                continue
                        else:
                            if rr[idx] / tmp >= 0.6 and s_tmp > 0 and rr[idx] / s_tmp >= 1.3:
                                detectsym2[idx] = 'N'
                                if divba[idx] < 1.4:
                                    flag_N = True
                                    continue
                            else:
                                if rr[idx] >= 300:
                                    detectsym2[idx] = 'N'
                                else:
                                    detectsym2[idx] = 'V'
                    else:
                        if rr[idx] >= 300:
                            detectsym2[idx] = 'N'
                        else:
                            detectsym2[idx] = 'V'
            if detectsym2[idx - 1] == 'V' and cnt_tmp>0:
                tmp = np.mean(rr[idx - cnt_tmp:idx-1])
                tmp_cnt1=cnt
                cnt = 0
                if idx==189:
                    m=0
                while(rr[idx]/tmp < 0.6):
                        idx+=1
                        cnt+=1
                        if idx >= len(divba) - 1:
                            break
                s_tmp = np.mean(rr[idx-cnt-1:idx])
                if cnt ==0:
                    # cnt_tmp=tmp_cnt1
                    # detectsym2[idx]='N'
                    # idx -= 1
                    # cnt -= 1
                    flag_N = True
                    if idx >= len(divba) - 1:
                        break
                    else:
                        continue
                else:
                    detectsym2[idx - cnt:idx] = 'V'
                    # detectsym2[idx] = 'N'
                    # idx-=1
                    # cnt -= 1
                    flag_N = True
                    if idx >= len(divba) - 1:
                        break
                    else:
                        continue



            # if detectsym2[idx-1]=='D':
            #     tmp = np.mean(rr[idx-cnt_tmp:idx])
            #     idx += 3
            #     cnt=0
            #     while(rr[idx]/tmp < 0.8):
            #         idx+=1
            #         cnt+=1
            #     detectsym2[idx-cnt:idx]='D'
            # else:
            #     idx += 3
            #     cnt = 0
            #     while (divba[idx] > 0.6 and (divba[idx + 1] - divba[idx] < 0.3) and divba[idx]<1.4):
            #         idx += 1
            #         cnt += 1
            #     detectsym2[idx - cnt:idx + 1] = 'N'
            #     if rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[1:idx + 1]) <= 1.2:
            #         detectsym2[idx + 1] = 'N'
            #         if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 2]) >= 0.8 and rr[idx + 2] / np.mean(
            #                 rr[idx - cnt:idx + 2]) <= 1.2:
            #             detectsym2[idx + 2] = 'N'
            #         else:
            #             detectsym2[idx + 2] = 'D'
            #     elif rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) > 1.2:
            #         detectsym2[idx + 1] = 'N'
            #         if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(
            #                 rr[idx - cnt:idx + 1]) <= 1.2:
            #             detectsym2[idx + 1:idx + 3] = 'N'
            #         else:
            #             detectsym2[idx + 2] = 'D'
        idx+=1
        cnt+=1
        if idx >= len(divba)-1:
            break
    s=0

    # if detectsym2[idx+2]=='D':
    #     tmp = np.mean(rr[idx-cnt:idx+2])
    #     idx += 3
    #     cnt=0
    #     while(rr[idx]/tmp < 0.8):
    #         idx+=1
    #         cnt+=1
    #     detectsym2[idx-cnt:idx]='D'
    # else:
    #     idx += 3
    #     cnt = 0
    #     while (divba[idx] > 0.6 and (divba[idx + 1] - divba[idx] < 0.3) and divba[idx]<1.4):
    #         idx += 1
    #         cnt += 1
    #     detectsym2[idx - cnt:idx + 1] = 'N'
    #     if rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[1:idx + 1]) <= 1.2:
    #         detectsym2[idx + 1] = 'N'
    #         if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 2]) >= 0.8 and rr[idx + 2] / np.mean(
    #                 rr[idx - cnt:idx + 2]) <= 1.2:
    #             detectsym2[idx + 2] = 'N'
    #         else:
    #             detectsym2[idx + 2] = 'D'
    #     elif rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) > 1.2:
    #         detectsym2[idx + 1] = 'N'
    #         if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(
    #                 rr[idx - cnt:idx + 1]) <= 1.2:
    #             detectsym2[idx + 1:idx + 3] = 'N'
    #         else:
    #             detectsym2[idx + 2] = 'D'
    #
    #
    #
    # if detectsym2[idx+2]=='D':
    #     tmp = np.mean(rr[idx-cnt:idx+2])
    #     idx += 3
    #     cnt=0
    #     while(rr[idx]/tmp < 0.8):
    #         idx+=1
    #         cnt+=1
    #     detectsym2[idx-cnt:idx]='D'
    #
    #
    # cnt_tmp= cnt
    # cnt=0
    # while (divba[idx]>0.6 and (divba[idx+1]-divba[idx]<0.3) and divba[idx]<1.4):
    #     idx+=1
    #     cnt+=1
    # if cnt ==0:
    #     if detectsym2[idx-1]=='D':
    #         if rr[idx]/np.mean(rr[idx - cnt_tmp-1:idx]) <= 1.2:
    #             detectsym2[idx]='D'
    #         else:
    #             detectsym2[idx] = 'N'
    #     else:
    #         if rr[idx]/np.mean(rr[idx - cnt_tmp-1:idx]) > 1.2:
    #             detectsym2[idx]='N'
    #         else:
    #             detectsym2[idx] = 'D'
    # else:
    #     detectsym2[idx - cnt:idx + 1] = 'N'
    #     if rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 1] / np.mean(rr[1:idx + 1]) <= 1.2:
    #         detectsym2[idx + 1] = 'N'
    #         if rr[idx+2]/np.mean(rr[idx-cnt:idx+2]) >= 0.8 and rr[idx+2]/np.mean(rr[idx-cnt:idx+2]) <= 1.2:
    #             detectsym2[idx+2]='N'
    #         else:
    #             detectsym2[idx + 2] = 'D'
    #     elif rr[idx + 1] / np.mean(rr[idx-cnt:idx + 1]) > 1.2:
    #         detectsym2[idx + 1] = 'N'
    #         if rr[idx + 2] / np.mean(rr[idx-cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[idx-cnt:idx + 1]) <= 1.2:
    #             detectsym2[idx + 1:idx + 3] = 'N'
    #         else:
    #             detectsym2[idx + 2] = 'D'
    #
    # idx += 3
    # cnt = 0
    # while (divba[idx] > 0.6 and (divba[idx + 1] - divba[idx] < 0.3) and divba[idx]<1.4):
    #     idx += 1
    #     cnt += 1
    # detectsym2[idx - cnt:idx + 1] = 'N'
    # if rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[1:idx + 1]) <= 1.2:
    #     detectsym2[idx + 1] = 'N'
    #     if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 2]) >= 0.8 and rr[idx + 2] / np.mean(rr[idx - cnt:idx + 2]) <= 1.2:
    #         detectsym2[idx + 2] = 'N'
    #     else:
    #         detectsym2[idx + 2] = 'D'
    # elif rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) > 1.2:
    #     detectsym2[idx + 1] = 'N'
    #     if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[idx - cnt:idx + 1]) <= 1.2:
    #         detectsym2[idx + 1:idx + 3] = 'N'
    #     else:
    #         detectsym2[idx + 2] = 'D'
    #
    # idx += 3
    # cnt = 0
    # fl = False
    # while (divba[idx] > 0.6 and (divba[idx + 1] - divba[idx] < 0.3) and divba[idx]<1.4):
    #     idx += 1
    #     cnt += 1
    #     if idx >= len(divba)-1:
    #         fl=True
    #         break
    # detectsym2[idx - cnt:idx + 1] = 'N'
    # if fl==False:
    #     if rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(rr[1:idx + 1]) <= 1.2:
    #         detectsym2[idx + 1] = 'N'
    #         if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 2]) >= 0.8 and rr[idx + 2] / np.mean(
    #                 rr[idx - cnt:idx + 2]) <= 1.2:
    #             detectsym2[idx + 2] = 'N'
    #         else:
    #             detectsym2[idx + 2] = 'D'
    #     elif rr[idx + 1] / np.mean(rr[idx - cnt:idx + 1]) > 1.2:
    #         detectsym2[idx + 1] = 'N'
    #         if rr[idx + 2] / np.mean(rr[idx - cnt:idx + 1]) >= 0.8 and rr[idx + 2] / np.mean(
    #                 rr[idx - cnt:idx + 1]) <= 1.2:
    #             detectsym2[idx + 1:idx + 3] = 'N'
    #         else:
    #             detectsym2[idx + 2] = 'D'


    #turn 2

    minlocal1 = []
    minlocal2 = []
    QS2 = []
    QS5 = []
    QS8 = []
    diff_f2 = [0]
    diff_a2 = []
    diff_f5 = [0]
    diff_a5 = []
    diff_f8 = [0]
    diff_a8 = []
    diff_fm = [0]
    diff_am = []
    diff_tb = []
    # print(namefile)
    for i in range(len(bts)):
        # print(i)
        if ecg[bts[i]] >= 0:
            d = 1
        else:
            d = -1

        if bts[i] < 30:
            s1 = 0 + np.argmin(d * ecg[0:bts[i]])
        else:
            s1 = bts[i] - 30 + np.argmin(d * ecg[bts[i] - 30:bts[i]])

        if bts[i] > len(ecg) - 30:
            s2 = bts[i] + np.argmin(d * ecg[bts[i]:])
        else:
            s2 = bts[i] + np.argmin(d * ecg[bts[i]:bts[i] + 30])
        if ecg[s1] * ecg[s2] >= 0:
            if abs(ecg[s1]) > abs(ecg[s2]):
                s12 = s1
            else:
                s12 = s2
        else:
            if ecg[s1] * ecg[bts[i]] < 0:
                s12 = s1
            else:
                s12 = s2

        diff_tb.append(round(abs(ecg[bts][i] - ecg[s12]), 6))

        cnt = 0
        f = True
        while (ecg[bts[i] - cnt] * ecg[bts[i]] > 0):
            cnt += 1
            if bts[i] - cnt < 0 or cnt > 30:
                minlocal1.append(bts[i] - cnt + 1)
                f = False
                break
        if f:
            minlocal1.append(bts[i] - cnt)
        cnt = 0
        f = True
        while (ecg[bts[i] + cnt] * ecg[bts[i]] > 0):
            cnt += 1
            if bts[i] + cnt > len(ecg) - 1 or cnt > 30:
                minlocal2.append(bts[i] + cnt - 1)
                f = False
                break
        if f:
            minlocal2.append(bts[i] + cnt)
        # print(i)
        sRS = 0.8 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.8 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS2.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
                np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.5 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.5 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS5.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
                np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.2 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.2 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS8.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
                np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
        if i > 0 and i < len(bts) - 1:
            diff_f2.append(round(abs((QS2[i] - QS2[i - 1]) / QS2[i - 1]), 6))
            diff_a2.append(round(abs((QS2[i - 1] - QS2[i]) / QS2[i]), 6))
            diff_f5.append(round(abs((QS5[i] - QS5[i - 1]) / QS5[i - 1]), 6))
            diff_a5.append(round(abs((QS5[i - 1] - QS5[i]) / QS5[i]), 6))
            diff_f8.append(round(abs((QS8[i] - QS8[i - 1]) / QS8[i - 1]), 6))
            diff_a8.append(round(abs((QS8[i - 1] - QS8[i]) / QS8[i]), 6))
            diff_fm.append(round((diff_f2[i] + diff_f5[i] + diff_f8[i]) / 3, 6))
            diff_am.append(round((diff_a2[i - 1] + diff_a5[i - 1] + diff_a8[i - 1]) / 3, 6))
        if i == len(bts) - 1:
            diff_a2.append(0)
            diff_a5.append(0)
            diff_a8.append(0)
            diff_am.append(0)

    for i in range(2,len(detectsym2)-2):
        if detectsym2[i]=='N' and divba[i] < 0.85:
            if (ecg[bts[i]] * ecg[bts[i - 1]] < 0 and ecg[bts[i]] * ecg[bts[i + 1]] < 0) or (diff_fm[i] > 1.3 and diff_am[i] > 1.3):
                detectsym2[i]='V'
    detect = detectsym2.copy()
    QS2 = np.array(QS2)
    QS5 = np.array(QS5)
    QS8 = np.array(QS8)
    idxn = 1
    while(True):
        cnt = 0
        while(detectsym2[idxn]=='N'):
            idxn+=1
            cnt+=1
            if idxn >= len(detectsym2) - 1:
                break
        if cnt==0:
            tmp2 = QS2[idxn]
            tmp5 = QS5[idxn]
            tmp8 = QS8[idxn]
        else:
            tmp2 = np.mean(QS2[idxn-cnt:idxn])
            tmp5 = np.mean(QS5[idxn-cnt:idxn])
            tmp8 = np.mean(QS8[idxn-cnt:idxn])
        if idxn >= len(detectsym2) - 1:
            break
        while(detectsym2[idxn]!='N'):

            tmp = (abs(QS2[idxn] - tmp2) / tmp2 + abs(QS5[idxn] - tmp5) / tmp5 + abs(QS8[idxn] - tmp8) / tmp8) / 2
            if tmp<0.2 and (ecg[bts[idxn]]*ecg[bts[idxn-1]]>0 or ecg[bts[idxn]]*ecg[bts[idxn+1]]>0):
                detectsym2[idxn]='S'
            idxn+=1
            if idxn >= len(detectsym2) - 1:
                break
        if idxn>=len(detectsym2)-1:
            break

    s=0

    wfdb.wrann(record_name=str(namefile),
               extension='atr',
               sample=np.asarray(bts),
               symbol=np.asarray(sym),
               fs=360,
               write_dir='PTK')

    # wfdb.wrann(record_name=str(namefile),
    #            extension='txt',
    #            sample=np.asarray(btss),
    #            symbol=np.asarray(detectsym2),
    #            fs=360,
    #            write_dir='PTK')

    wfdb.wrann(record_name=str(namefile),
               extension='txt',
               sample=np.asarray(bts),
               symbol=np.asarray(detectsym2),
               fs=360,
               write_dir='PTK')

def ann_ptk(DIR, namefile):
    record = wfdb.rdrecord(DIR + str(namefile))
    ecg_o = record.p_signal[:, 0]
    d_ptk_bts, symbol_ptk = pan_tompkin(ecg_o, fs=record.fs)

    ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
    bts = np.zeros_like(beats)
    print(namefile)
    for i in range(len(beats)):
        if ecg[beats[i]] < 0:
            # print(i, ' ', 1)
            bts[i] = np.argmax(
                -(ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
        else:
            bts[i] = np.argmax(
                (ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
            # print(i, ' ', 2)
    l_btss = len(d_ptk_bts)
    top = 0
    bot = 0
    if d_ptk_bts[0] - 30 < 0:
        l_btss -= 1
        symbol_ptk = symbol_ptk[1:]
        top = 1
    if d_ptk_bts[-1] + 30 > len(ecg):
        l_btss -= 1
        symbol_ptk = symbol_ptk[:-1]
        bot = 1
    btss = np.zeros(l_btss, dtype='int64')

    for j in range(len(d_ptk_bts) - bot - top):
        btss[j] = np.argmax(abs(ecg[d_ptk_bts[j + top] - 30:d_ptk_bts[j + top] + 30])) + d_ptk_bts[j + top] - 30


    sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
    sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'S'
    sym[(sym == 'V') | (sym == 'e')] = 'V'
    sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'

    symidx = np.zeros(len(sym))
    nondata = sym.copy()
    nondata[:] = ' '
    symidx[sym != 'N'] = 1

    detectsym2 = np.array(symbol_ptk.copy())
    detectsym2[:2]='N'
    detectsym2[-2:]='N'


    df = np.diff(np.array(btss))
    if df[0]<200:
        df[0]=200
    rr = [0]
    rr.extend(df)
    rr = np.array(rr)
    divba = np.zeros(len(btss))
    divba[1:-1] = df[:-1] / df[1:]

    # turn 1 detect SV
    flag_N = False

    idx=1
    cnt = 0

    while (divba[idx] > 0.6 and (abs(divba[idx + 1] - divba[idx]) < 0.3) and divba[idx] < 1.4):
        idx += 1
        cnt += 1

    detectsym2[idx - cnt:idx + 1] = 'N'
    idx += 1
    cnt += 1
    s_tmp = -1

    while (True):
        if flag_N:

            cnt = 0
            tmp_while = rr[idx]
            while (divba[idx] > 0.6 and (abs(divba[idx + 1] - divba[idx]) < 0.3) and divba[idx] < 1.4 and rr[
                idx + 1] / tmp_while > 0.6):
                # print(idx)
                idx += 1
                cnt += 1
                if idx >= len(divba) - 1:
                    break
                tmp_while = np.mean(rr[idx - cnt:idx])
            # if cnt==1:
            #     cnt=0
            # else:
            detectsym2[idx - cnt:idx + 1] = 'N'

            # idx+=1
            flag_N = False
        else:
            cnt_tmp = cnt
            if detectsym2[idx - 1] == 'N' and cnt_tmp > 0:
                if idx == 1837:
                    m = 0

                ik = idx-cnt-1
                if ik<1:
                    ik=1
                while(detectsym2[ik]!='N'):
                        ik-=1
                        if ik<1:
                            ik=1
                            break
                for k in range(cnt):
                    if rr[idx-k - 1]/rr[ik]>2:
                        rr[idx-k - 1]=rr[ik]


                tmp = np.mean(rr[idx - cnt_tmp:idx])
                if rr[idx] / tmp >= 0.7:
                    detectsym2[idx] = 'N'
                    if rr[idx] / tmp > 2:
                        rr[idx] = tmp
                    idx += 1
                    cnt += 1
                    tmp = np.mean(rr[idx - cnt:idx])
                    if idx >= len(divba) - 1:
                        break

                    if rr[idx] / tmp >= 0.7:
                        detectsym2[idx] = 'N'
                        if rr[idx] / tmp > 2:
                            rr[idx] = tmp
                        if divba[idx] < 1.4:
                            flag_N = True
                            continue

                    else:
                        if rr[idx] / tmp >= 0.6 and s_tmp > 0 and rr[idx] / s_tmp >= 1.3:
                            detectsym2[idx] = 'N'
                            if divba[idx] < 1.4:
                                flag_N = True
                                continue
                        else:
                            if rr[idx] >= 300:
                                detectsym2[idx] = 'N'
                            else:
                                detectsym2[idx] = 'V'
                    # else den noi de so sanh RR với mean{RR} truoc
                else:
                    if rr[idx] / tmp >= 0.6 and s_tmp > 0 and rr[idx] / s_tmp >= 1.3:
                        detectsym2[idx] = 'N'
                        # if rr[idx] / tmp < 1.3:
                        #     tmp = np.mean(rr[idx - cnt_tmp:idx])
                        idx += 1
                        cnt += 1
                        tmp = np.mean(rr[idx - cnt:idx])
                        if idx >= len(divba) - 1:
                            break
                        # cnt_tmp = cnt

                        if rr[idx] / tmp >= 0.7:
                            detectsym2[idx] = 'N'
                            if rr[idx] / tmp > 2:
                                rr[idx] = tmp
                            if divba[idx] < 1.4:
                                flag_N = True
                                continue
                        else:
                            if rr[idx] / tmp >= 0.6 and s_tmp > 0 and rr[idx] / s_tmp >= 1.3:
                                detectsym2[idx] = 'N'
                                if divba[idx] < 1.4:
                                    flag_N = True
                                    continue
                            else:
                                if rr[idx] >= 300:
                                    detectsym2[idx] = 'N'
                                else:
                                    detectsym2[idx] = 'V'
                    else:
                        if rr[idx] >= 300:
                            detectsym2[idx] = 'N'
                        else:
                            detectsym2[idx] = 'V'
            if detectsym2[idx - 1] == 'V' and cnt_tmp > 0:
                tmp = np.mean(rr[idx - cnt_tmp:idx - 1])
                tmp_cnt1 = cnt
                cnt = 0
                if idx == 1839:
                    m = 0
                while (rr[idx] / tmp < 0.6):
                    idx += 1
                    cnt += 1
                    if idx >= len(divba) - 1:
                        break
                s_tmp = np.mean(rr[idx - cnt - 1:idx])
                if cnt == 0:
                    # cnt_tmp=tmp_cnt1
                    # detectsym2[idx]='N'
                    # idx -= 1
                    # cnt -= 1
                    flag_N = True
                    if idx >= len(divba) - 1:
                        break
                    else:
                        continue
                else:
                    detectsym2[idx - cnt:idx] = 'V'
                    # detectsym2[idx] = 'N'
                    # idx-=1
                    # cnt -= 1
                    flag_N = True
                    if idx >= len(divba) - 1:
                        break
                    else:
                        continue

        idx += 1
        cnt += 1
        if idx >= len(divba) - 1:
            break

    # turn 2 detect SV

    minlocal1 = []
    minlocal2 = []
    QS2 = []
    QS5 = []
    QS8 = []
    diff_f2 = [0]
    diff_a2 = []
    diff_f5 = [0]
    diff_a5 = []
    diff_f8 = [0]
    diff_a8 = []
    diff_fm = [0]
    diff_am = []
    diff_tb = []
    # print(namefile)
    for i in range(len(btss)):
        # print(i)
        if ecg[btss[i]] >= 0:
            d = 1
        else:
            d = -1

        if btss[i] < 30:
            s1 = 0 + np.argmin(d * ecg[0:btss[i]])
        else:
            s1 = btss[i] - 30 + np.argmin(d * ecg[btss[i] - 30:btss[i]])

        if btss[i] > len(ecg) - 30:
            s2 = btss[i] + np.argmin(d * ecg[btss[i]:])
        else:
            s2 = btss[i] + np.argmin(d * ecg[btss[i]:btss[i] + 30])
        if ecg[s1] * ecg[s2] >= 0:
            if abs(ecg[s1]) > abs(ecg[s2]):
                s12 = s1
            else:
                s12 = s2
        else:
            if ecg[s1] * ecg[btss[i]] < 0:
                s12 = s1
            else:
                s12 = s2

        diff_tb.append(round(abs(ecg[btss][i] - ecg[s12]), 6))

        cnt = 0
        f = True
        while (ecg[btss[i] - cnt] * ecg[btss[i]] > 0):
            cnt += 1
            if btss[i] - cnt < 0 or cnt > 30:
                minlocal1.append(btss[i] - cnt + 1)
                f = False
                break
        if f:
            minlocal1.append(btss[i] - cnt)
        cnt = 0
        f = True
        while (ecg[btss[i] + cnt] * ecg[btss[i]] > 0):
            cnt += 1
            if btss[i] + cnt > len(ecg) - 1 or cnt > 30:
                minlocal2.append(btss[i] + cnt - 1)
                f = False
                break
        if f:
            minlocal2.append(btss[i] + cnt)
        # print(i)
        sRS = 0.8 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.8 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS2.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.5 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.5 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS5.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.2 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.2 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS8.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
        if i > 0 and i < len(btss) - 1:
            diff_f2.append(round(abs((QS2[i] - QS2[i - 1]) / QS2[i - 1]), 6))
            diff_a2.append(round(abs((QS2[i - 1] - QS2[i]) / QS2[i]), 6))
            diff_f5.append(round(abs((QS5[i] - QS5[i - 1]) / QS5[i - 1]), 6))
            diff_a5.append(round(abs((QS5[i - 1] - QS5[i]) / QS5[i]), 6))
            diff_f8.append(round(abs((QS8[i] - QS8[i - 1]) / QS8[i - 1]), 6))
            diff_a8.append(round(abs((QS8[i - 1] - QS8[i]) / QS8[i]), 6))
            diff_fm.append(round((diff_f2[i] + diff_f5[i] + diff_f8[i]) / 3, 6))
            diff_am.append(round((diff_a2[i - 1] + diff_a5[i - 1] + diff_a8[i - 1]) / 3, 6))
        if i == len(btss) - 1:
            diff_a2.append(0)
            diff_a5.append(0)
            diff_a8.append(0)
            diff_am.append(0)

    for i in range(2, len(detectsym2) - 2):
        if detectsym2[i] == 'N' and divba[i] < 0.85:
            if (ecg[btss[i]] * ecg[btss[i - 1]] < 0 and ecg[btss[i]] * ecg[btss[i + 1]] < 0) or (
                    diff_fm[i] > 1.3 and diff_am[i] > 1.3):
                detectsym2[i] = 'V'
    s=0

    # classify S and V
    QS2 = np.array(QS2)
    QS5 = np.array(QS5)
    QS8 = np.array(QS8)
    idxn = 1
    while(True):
        cnt = 0
        while(detectsym2[idxn]=='N'):
            idxn+=1
            cnt+=1
            if idxn >= len(detectsym2) - 1:
                break
        if cnt==0:
            tmp2 = QS2[idxn]
            tmp5 = QS5[idxn]
            tmp8 = QS8[idxn]
        else:
            tmp2 = np.mean(QS2[idxn-cnt:idxn])
            tmp5 = np.mean(QS5[idxn-cnt:idxn])
            tmp8 = np.mean(QS8[idxn-cnt:idxn])
        if idxn >= len(detectsym2) - 1:
            break
        while(detectsym2[idxn]!='N'):

            tmp = (abs(QS2[idxn] - tmp2) / tmp2 + abs(QS5[idxn] - tmp5) / tmp5 + abs(QS8[idxn] - tmp8) / tmp8) / 2
            if tmp<0.2 and (ecg[btss[idxn]]*ecg[btss[idxn-1]]>0 or ecg[btss[idxn]]*ecg[btss[idxn+1]]>0):
                detectsym2[idxn]='S'
            idxn+=1
            if idxn >= len(detectsym2) - 1:
                break
        if idxn>=len(detectsym2)-1:
            break

    wfdb.wrann(record_name=str(namefile),
               extension='atr',
               sample=np.asarray(bts),
               symbol=np.asarray(sym),
               fs=360,
               write_dir='PTK')



    wfdb.wrann(record_name=str(namefile),
               extension='txt',
               sample=np.asarray(btss),
               symbol=np.asarray(detectsym2),
               fs=360,
               write_dir='PTK')

def ann_origin(DIR, namefile):
    record = wfdb.rdrecord(DIR + str(namefile))
    ecg_o = record.p_signal[:, 0]
    d_ptk_bts, symbol_ptk = pan_tompkin(ecg_o, fs=record.fs)

    ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
    bts = np.zeros_like(beats)
    print(namefile)
    for i in range(len(beats)):
        if ecg[beats[i]] < 0:
            # print(i, ' ', 1)
            bts[i] = np.argmax(
                -(ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
        else:
            bts[i] = np.argmax(
                (ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
            # print(i, ' ', 2)
    l_btss = len(d_ptk_bts)
    top = 0
    bot = 0
    if d_ptk_bts[0] - 30 < 0:
        l_btss -= 1
        symbol_ptk = symbol_ptk[1:]
        top = 1
    if d_ptk_bts[-1] + 30 > len(ecg):
        l_btss -= 1
        symbol_ptk = symbol_ptk[:-1]
        bot = 1
    btss = np.zeros(l_btss, dtype='int64')

    for j in range(len(d_ptk_bts) - bot - top):
        btss[j] = np.argmax(abs(ecg[d_ptk_bts[j + top] - 30:d_ptk_bts[j + top] + 30])) + d_ptk_bts[j + top] - 30

    sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
    sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'S'
    sym[(sym == 'V') | (sym == 'e')] = 'V'
    sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'
    fill_colorV = np.where(sym == 'V')[0] + 2
    fill_colorS = np.where(sym == 'S')[0] + 2
    symidx = np.zeros(len(sym))
    nondata = sym.copy()
    nondata[:] = ' '
    symidx[sym != 'N'] = 1

    # detectsym2 = symbol_ptk.copy()
    #
    # # detectsym2[2:len(detectsym2)-1] = ' '
    #
    # df = np.diff(np.array(btss))
    # divba = np.zeros(len(btss))
    # divba[1:-1] = df[:-1] / df[1:]
    #
    # for i in range(2, len(btss) - 2):
    #     if divba[i - 1] >= 1.2:
    #         if divba[i] < 0.7:
    #             if divba[i - 1] >= 2:
    #                 detectsym2[i] = 'V'
    #             elif divba[i + 1] >= 1:
    #                 detectsym2[i] = 'V'
    #             else:
    #                 detectsym2[i] = 'N'
    #         else:
    #             detectsym2[i] = 'N'
    #     else:
    #         detectsym2[i] = 'N'

    detectsym2 = sym.copy()
    df = np.diff(np.array(bts))
    divba = np.zeros(len(bts))
    divba[1:-1] = df[:-1] / df[1:]

    for i in range(2, len(bts) - 2):
        if divba[i - 1] >= 1.2:
            if divba[i] < 0.7:
                if divba[i - 1] >= 2:
                    detectsym2[i] = 'V'
                elif divba[i + 1] >= 1:
                    detectsym2[i] = 'V'
                else:
                    detectsym2[i] = 'N'
            else:
                detectsym2[i] = 'N'
        else:
            detectsym2[i] = 'N'
    time = bts/FREQUENCY_SAMPLING
    # sym1 = sym.copy()
    # sym1[sym=='S']='V'
    # wfdb.wrann(record_name=str(namefile),
    #            extension='atr',
    #            sample=np.asarray(bts),
    #            symbol=np.asarray(sym),
    #            fs=360,
    #            write_dir='PTK')
    #
    # # wfdb.wrann(record_name=str(namefile),
    # #            extension='txt',
    # #            sample=np.asarray(btss),
    # #            symbol=np.asarray(detectsym2),
    # #            fs=360,
    # #            write_dir='PTK')
    #
    # wfdb.wrann(record_name=str(namefile),
    #            extension='txt',
    #            sample=np.asarray(bts),
    #            symbol=np.asarray(detectsym2),
    #            fs=360,
    #            write_dir='PTK')

    sym2 = sym.copy()

    idx2 = 0
    # while (idx2 < len(sym)):
    #     cnt = 0
    #     while (sym2[idx2] != 'N'):
    #         cnt += 1
    #         idx2 += 1
    #         if idx2 >= len(sym):
    #             break
    #     if cnt == 1:
    #         # sym2[idx2 - 1] = 'V'
    #         sym2[idx2 - 1] = sym[idx2-1]
    #     else:
    #         sym2[idx2 - cnt:idx2] = 'N'
    #     idx2 += 1

    # minlocal1 = []
    # minlocal2 = []
    # QRv = []
    # QRs = []
    # RSv = []
    # RSs = []
    # QSv = []
    # QSs = []
    # QS2 = []
    # QS5 = []
    # QS8 = []
    #
    # print(namefile)
    # for i in range(len(btss)):
    #     # print(i)
    #     if ecg[btss[i]] >= 0:
    #         d = 1
    #     else:
    #         d = -1
    #     if btss[i] < 30:
    #         print(i)
    #         minlocal1.append(0 + np.argmin(d * ecg[0:btss[i]]))
    #     else:
    #         minlocal1.append(btss[i] - 30 + np.argmin(d * ecg[btss[i] - 30:btss[i]]))
    #
    #     if btss[i] > len(ecg) - 30:
    #         minlocal2.append(btss[i] + np.argmin(d * ecg[btss[i]:]))
    #     else:
    #         minlocal2.append(btss[i] + np.argmin(d * ecg[btss[i]:btss[i] + 30]))
    #
    #     QSv.append(ecg[minlocal2[i]] - ecg[minlocal1[i]])
    #     QSs.append(minlocal2[i] - minlocal1[i])
    #
    #     QRv.append(ecg[btss[i]] - ecg[minlocal1[i]])
    #     QRs.append(btss[i] - minlocal1[i])
    #
    #     RSv.append(ecg[btss[i]] - ecg[minlocal2[i]])
    #     RSs.append(minlocal2[i] - btss[i])
    #
    #     sRS = 0.2 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
    #     sQR = 0.2 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
    #     QS2.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
    #             np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
    #
    #     sRS = 0.5 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
    #     sQR = 0.5 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
    #     QS5.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
    #             np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
    #
    #     sRS = 0.8 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
    #     sQR = 0.8 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
    #     QS8.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
    #             np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
    #
    # diff_f = []
    # diff_a = []
    # diff_f.append(np.mean(np.zeros(9)))
    # diff_a.append(round(np.mean(np.abs(np.array(
    #     [(QSv[1] / QSv[2]) ** 2, (QSs[1] / QSs[2]) ** 2, (QRs[1] / QRs[2]) ** 2, (RSs[1] / RSs[2]) ** 2,
    #      (QS2[1] / QS2[2]) ** 2, (QS5[1] / QS5[2]) ** 2,
    #      (QS8[1] / QS8[2]) ** 2]))), 5))
    #
    # for i in range(1, len(QSs) - 1):
    #     diff_f.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i - 1] / QSv[i]) ** 2, (QSs[i - 1] / QSs[i]) ** 2, (QRs[i - 1] / QRs[i]) ** 2,
    #          (RSs[i - 1] / RSs[i]) ** 2, (QS2[i - 1] / QS2[i]) ** 2,
    #          (QS5[i - 1] / QS5[i]) ** 2, (QS8[i - 1] / QS8[i]) ** 2]))), 5))
    #     diff_a.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i] / QSv[i + 1]) ** 2, (QSs[i] / QSs[i + 1]) ** 2, (QRs[i] / QRs[i + 1]) ** 2,
    #          (RSs[i] / RSs[i + 1]) ** 2, (QS2[i] / QS2[i + 1]) ** 2,
    #          (QS5[i] / QS5[i + 1]) ** 2, (QS8[i] / QS8[i + 1]) ** 2]))), 5))
    # diff_f.append(round(np.mean(np.abs(np.array(
    #     [(QSv[-2] / QSv[-1]) ** 2, (QSs[-2] / QSs[-1]) ** 2, (QRs[-2] / QRs[-1]) ** 2, (RSs[-2] / RSs[-1]) ** 2,
    #      (QS2[-2] / QS2[-1]) ** 2,
    #      (QS5[-2] / QS5[-1]) ** 2, (QS8[-2] / QS8[-1]) ** 2]))), 5))
    # diff_a.append(np.mean(np.zeros(9)))

    # minlocal1 = []
    # minlocal2 = []
    # QRv = []
    # QRs = []
    # RSv = []
    # RSs = []
    # QSv = []
    # QSs = []
    # QS2 = []
    # QS5 = []
    # QS8 = []
    # print(namefile)
    # for i in range(len(bts)):
    #     # print(i)
    #     if ecg[bts[i]] >= 0:
    #         d = 1
    #     else:
    #         d = -1
    #     if bts[i] < 30:
    #         minlocal1.append(0 + np.argmin(d * ecg[0:bts[i]]))
    #     else:
    #         minlocal1.append(bts[i] - 30 + np.argmin(d * ecg[bts[i] - 30:bts[i]]))
    #
    #     if bts[i] > len(ecg) - 30:
    #         minlocal2.append(bts[i] + np.argmin(d * ecg[bts[i]:]))
    #     else:
    #         minlocal2.append(bts[i] + np.argmin(d * ecg[bts[i]:bts[i] + 30]))
    #
    #     QSv.append(ecg[minlocal2[i]] - ecg[minlocal1[i]])
    #     QSs.append(minlocal2[i] - minlocal1[i])
    #
    #     QRv.append(ecg[bts[i]] - ecg[minlocal1[i]])
    #     QRs.append(bts[i] - minlocal1[i])
    #
    #     RSv.append(ecg[bts[i]] - ecg[minlocal2[i]])
    #     RSs.append(minlocal2[i] - bts[i])
    #
    #     sRS = 0.2 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
    #     sQR = 0.2 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
    #     QS2.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
    #             np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
    #
    #     sRS = 0.5 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
    #     sQR = 0.5 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
    #     QS5.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
    #             np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
    #
    #     sRS = 0.8 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
    #     sQR = 0.8 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
    #     QS8.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
    #             np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))

    # diff_f = []
    # diff_a = []
    #
    # diff_f.append(np.mean(np.zeros(9)))
    # diff_a.append(round(np.mean(np.abs(np.array(
    #     [(QSv[1] / QSv[2]) ** 2, (QSs[1] / QSs[2]) ** 2, (QRs[1] / QRs[2]) ** 2, (RSs[1] / RSs[2]) ** 2,
    #      (QS2[1] / QS2[2]) ** 2, (QS5[1] / QS5[2]) ** 2,
    #      (QS8[1] / QS8[2]) ** 2]))), 5))
    #
    #
    # for i in range(1, len(QSs) - 1):
    #     diff_f.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i - 1] / QSv[i]) ** 2, (QSs[i - 1] / QSs[i]) ** 2, (QRs[i - 1] / QRs[i]) ** 2,
    #          (RSs[i - 1] / RSs[i]) ** 2, (QS2[i - 1] / QS2[i]) ** 2,
    #          (QS5[i - 1] / QS5[i]) ** 2, (QS8[i - 1] / QS8[i]) ** 2]))), 5))
    #     diff_a.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i] / QSv[i + 1]) ** 2, (QSs[i] / QSs[i + 1]) ** 2, (QRs[i] / QRs[i + 1]) ** 2,
    #          (RSs[i] / RSs[i + 1]) ** 2, (QS2[i] / QS2[i + 1]) ** 2,
    #          (QS5[i] / QS5[i + 1]) ** 2, (QS8[i] / QS8[i + 1]) ** 2]))), 5))
    # diff_f.append(round(np.mean(np.abs(np.array(
    #     [(QSv[-2] / QSv[-1]) ** 2, (QSs[-2] / QSs[-1]) ** 2, (QRs[-2] / QRs[-1]) ** 2, (RSs[-2] / RSs[-1]) ** 2,
    #      (QS2[-2] / QS2[-1]) ** 2,
    #      (QS5[-2] / QS5[-1]) ** 2, (QS8[-2] / QS8[-1]) ** 2]))), 5))
    # diff_a.append(np.mean(np.zeros(9)))

    # for i in range(len(btss)):
    #     if detectsym2[i]=='V':
    #         if diff_a[i] <= 1.5:
    #         #     btss = 'N'
    #         # else:
    #             detectsym2[i] = 'S'

    # return sym, bts, time, divba, QSv, QSs, QRv, QRs, RSv, RSs, QS2, QS5, QS8, fill_colorV, fill_colorS, nondata, err, expt, expt2, err2

    minlocal1 = []
    minlocal2 = []
    QS2 = []
    QS5 = []
    QS8 = []
    diff_f2 = [0]
    diff_a2 = []
    diff_f5 = [0]
    diff_a5 = []
    diff_f8 = [0]
    diff_a8 = []
    diff_fm = [0]
    diff_am = []
    diff_tb = []
    # print(namefile)
    for i in range(len(bts)):
        # print(i)
        if ecg[bts[i]] >= 0:
            d = 1
        else:
            d = -1

        if bts[i] < 30:
            s1= 0 + np.argmin(d * ecg[0:bts[i]])
        else:
            s1 = bts[i] - 30 + np.argmin(d * ecg[bts[i] - 30:bts[i]])

        if bts[i] > len(ecg) - 30:
            s2 = bts[i] + np.argmin(d * ecg[bts[i]:])
        else:
           s2 = bts[i] + np.argmin(d * ecg[bts[i]:bts[i] + 30])
        if ecg[s1]*ecg[s2]>=0:
            if abs(ecg[s1]) > abs(ecg[s2]):
                s12 = s1
            else:
                s12 = s2
        else:
            if ecg[s1]*ecg[bts[i]]<0:
                s12 = s1
            else:
                s12 = s2

        diff_tb.append(round(abs(ecg[bts][i]-ecg[s12]),6))

        cnt = 0
        f = True
        while (ecg[bts[i] - cnt] * ecg[bts[i]] > 0):
            cnt += 1
            if bts[i] - cnt < 0 or cnt > 30:
                minlocal1.append(bts[i] - cnt + 1)
                f = False
                break
        if f:
            minlocal1.append(bts[i] - cnt)
        cnt = 0
        f = True
        while (ecg[bts[i] + cnt] * ecg[bts[i]] > 0):
            cnt += 1
            if bts[i] + cnt > len(ecg)-1 or cnt > 30:
                minlocal2.append(bts[i] + cnt - 1)
                f = False
                break
        if f:
            minlocal2.append(bts[i] + cnt)
        # print(i)
        sRS = 0.8 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.8 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS2.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
                np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.5 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.5 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS5.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
                np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.2 * (ecg[bts[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.2 * (ecg[bts[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS8.append(np.argmin(abs(ecg[bts[i]:minlocal2[i] + 1] - sRS)) + bts[i] - (
                np.argmin(abs(ecg[minlocal1[i]:bts[i] + 1] - sQR)) + minlocal1[i]))
        if i>0 and i < len(bts) - 1:
            diff_f2.append(round(abs((QS2[i]-QS2[i-1])/QS2[i-1]),6))
            diff_a2.append(round(abs((QS2[i-1]-QS2[i])/QS2[i]),6))
            diff_f5.append(round(abs((QS5[i] - QS5[i - 1])/QS5[i - 1]),6))
            diff_a5.append(round(abs((QS5[i - 1] - QS5[i])/QS5[i]),6))
            diff_f8.append(round(abs((QS8[i] - QS8[i - 1])/QS8[i - 1]),6))
            diff_a8.append(round(abs((QS8[i - 1] - QS8[i])/QS8[i]),6))
            diff_fm.append(round((diff_f2[i]+diff_f5[i]+diff_f8[i])/3,6))
            diff_am.append(round((diff_a2[i-1]+diff_a5[i-1]+diff_a8[i-1])/3,6))
        if i==len(bts)-1:
            diff_a2.append(0)
            diff_a5.append(0)
            diff_a8.append(0)
            diff_am.append(0)



    s=0
    # diff_f = []
    # diff_a = []
    #
    # diff_f.append(np.mean(np.zeros(9)))
    # diff_a.append(round(np.mean(np.abs(np.array(
    #     [(QSv[1] / QSv[2]) ** 2, (QSs[1] / QSs[2]) ** 2, (QRs[1] / QRs[2]) ** 2, (RSs[1] / RSs[2]) ** 2,
    #      (QS2[1] / QS2[2]) ** 2, (QS5[1] / QS5[2]) ** 2,
    #      (QS8[1] / QS8[2]) ** 2]))), 5))
    #
    #
    # for i in range(1, len(QSs) - 1):
    #     diff_f.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i - 1] / QSv[i]) ** 2, (QSs[i - 1] / QSs[i]) ** 2, (QRs[i - 1] / QRs[i]) ** 2,
    #          (RSs[i - 1] / RSs[i]) ** 2, (QS2[i - 1] / QS2[i]) ** 2,
    #          (QS5[i - 1] / QS5[i]) ** 2, (QS8[i - 1] / QS8[i]) ** 2]))), 5))
    #     diff_a.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i] / QSv[i + 1]) ** 2, (QSs[i] / QSs[i + 1]) ** 2, (QRs[i] / QRs[i + 1]) ** 2,
    #          (RSs[i] / RSs[i + 1]) ** 2, (QS2[i] / QS2[i + 1]) ** 2,
    #          (QS5[i] / QS5[i + 1]) ** 2, (QS8[i] / QS8[i + 1]) ** 2]))), 5))
    # diff_f.append(round(np.mean(np.abs(np.array(
    #     [(QSv[-2] / QSv[-1]) ** 2, (QSs[-2] / QSs[-1]) ** 2, (QRs[-2] / QRs[-1]) ** 2, (RSs[-2] / RSs[-1]) ** 2,
    #      (QS2[-2] / QS2[-1]) ** 2,
    #      (QS5[-2] / QS5[-1]) ** 2, (QS8[-2] / QS8[-1]) ** 2]))), 5))
    # diff_a.append(np.mean(np.zeros(9)))
    err = []
    err.extend(np.where(detectsym2 == 'V')[0][np.where(
        (sym2[np.where(detectsym2 == 'V')[0]] == 'N') | (sym2[np.where(detectsym2 == 'V')[0]] == 'O') | (
                sym2[np.where(detectsym2 == 'V')[0]] == 'F'))[0]] + 2)
    err.extend(np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0][np.where(
        (sym2[np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0]] == 'S') | (
            (sym2[np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0]] == 'V')))[0]] + 2)
    expt= []
    expt.extend(np.where(np.array(diff_fm)<0.15)[0] + 2)
    expt2 = []
    expt2.extend(np.where(np.array(diff_am) < 0.15)[0] + 2)
    err2 = []
    # return sym, bts, np.round(time,6), np.round(divba,6),np.round(ecg[bts],6),diff_tb, QS2, QS5, QS8, diff_f2, diff_a2,diff_f5, diff_a5, diff_f8, diff_a8, diff_fm, diff_am, fill_colorV, fill_colorS, nondata, err, expt, expt2, err2

    # for i in range(len(bts)):
    #     if detectsym2[i]=='V':
    #         if diff_a[i] <= 1.5:
    #         #     btss = 'N'
    #         # else:
    #             detectsym2[i] = 'S'
    # s = []
    for i in range(2,len(detectsym2)-2):
        if detectsym2[i]=='N' and divba[i] < 0.85:
            if (ecg[bts[i]] * ecg[bts[i - 1]] < 0 and ecg[bts[i]] * ecg[bts[i + 1]] < 0) or (diff_fm[i] > 1.3 and diff_am[i] > 1.3):
                detectsym2[i]='V'
                # s.append(i)
    for i in range(2, len(detectsym2)-2):
        if detectsym2[i]=='V':
            if (diff_fm[i]>=0.12 and diff_am[i]>=0.12) or max(diff_fm[i],diff_am[i])>=0.5 or (ecg[bts[i]] * ecg[bts[i - 1]] < 0 and ecg[bts[i]] * ecg[bts[i + 1]] < 0):
                detectsym2[i]='V'
            else:
                detectsym2[i]='S'


    wfdb.wrann(record_name=str(namefile),
               extension='atr',
               sample=np.asarray(bts),
               symbol=np.asarray(sym2),
               fs=360,
               write_dir='PTK')

    # wfdb.wrann(record_name=str(namefile),
    #            extension='txt',
    #            sample=np.asarray(btss),
    #            symbol=np.asarray(detectsym2),
    #            fs=360,
    #            write_dir='PTK')

    wfdb.wrann(record_name=str(namefile),
               extension='txt',
               sample=np.asarray(bts),
               symbol=np.asarray(detectsym2),
               fs=360,
               write_dir='PTK')

def ann_ptk_origin(DIR, namefile):
    record = wfdb.rdrecord(DIR + str(namefile))
    ecg_o = record.p_signal[:, 0]
    d_ptk_bts, symbol_ptk = pan_tompkin(ecg_o, fs=record.fs)

    ecg, beats, sym = preprocess_data(DIR + str(namefile) + '.hea')
    bts = np.zeros_like(beats)
    print(namefile)
    for i in range(len(beats)):
        if ecg[beats[i]] < 0:
            # print(i, ' ', 1)
            bts[i] = np.argmax(
                -(ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
        else:
            bts[i] = np.argmax(
                (ecg[beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN):beats[i] + int(FREQUENCY_SAMPLING * BXB_WIN)])) + \
                     beats[i] - int(FREQUENCY_SAMPLING * BXB_WIN)
            # print(i, ' ', 2)
    l_btss = len(d_ptk_bts)
    top = 0
    bot = 0
    if d_ptk_bts[0] - 30 < 0:
        l_btss -= 1
        symbol_ptk = symbol_ptk[1:]
        top = 1
    if d_ptk_bts[-1] + 30 > len(ecg):
        l_btss -= 1
        symbol_ptk = symbol_ptk[:-1]
        bot = 1
    btss = np.zeros(l_btss, dtype='int64')

    for j in range(len(d_ptk_bts) - bot - top):
        btss[j] = np.argmax(abs(ecg[d_ptk_bts[j + top] - 30:d_ptk_bts[j + top] + 30])) + d_ptk_bts[j + top] - 30

    sym[(sym == 'N') | (sym == 'L') | (sym == 'R') | (sym == 'e') | (sym == 'j')] = 'N'
    sym[(sym == 'A') | (sym == 'a') | (sym == 'J') | (sym == 'S')] = 'S'
    sym[(sym == 'V') | (sym == 'e')] = 'V'
    sym[(sym == 'f') | (sym == 'Q') | (sym == 'F')] = 'N'
    fill_colorV = np.where(sym == 'V')[0] + 2
    fill_colorS = np.where(sym == 'S')[0] + 2
    symidx = np.zeros(len(sym))
    nondata = sym.copy()
    nondata[:] = ' '
    symidx[sym != 'N'] = 1

    detectsym2 = symbol_ptk.copy()

    # detectsym2[2:len(detectsym2)-1] = ' '

    df = np.diff(np.array(btss))
    divba = np.zeros(len(btss))
    divba[1:-1] = df[:-1] / df[1:]

    for i in range(2, len(btss) - 2):
        if divba[i - 1] >= 1.2:
            if divba[i] < 0.7:
                if divba[i - 1] >= 2:
                    detectsym2[i] = 'V'
                elif divba[i + 1] >= 1:
                    detectsym2[i] = 'V'
                else:
                    detectsym2[i] = 'N'
            else:
                detectsym2[i] = 'N'
        else:
            detectsym2[i] = 'N'

    # detectsym2 = sym.copy()
    # df = np.diff(np.array(bts))
    # divba = np.zeros(len(bts))
    # divba[1:-1] = df[:-1] / df[1:]
    #
    # for i in range(2, len(bts) - 2):
    #     if divba[i - 1] >= 1.2:
    #         if divba[i] < 0.7:
    #             if divba[i - 1] >= 2:
    #                 detectsym2[i] = 'V'
    #             elif divba[i + 1] >= 1:
    #                 detectsym2[i] = 'V'
    #             else:
    #                 detectsym2[i] = 'N'
    #         else:
    #             detectsym2[i] = 'N'
    #     else:
    #         detectsym2[i] = 'N'



    sym2 = sym.copy()

    idx2 = 0
    # while (idx2 < len(sym)):
    #     cnt = 0
    #     while (sym2[idx2] != 'N'):
    #         cnt += 1
    #         idx2 += 1
    #         if idx2 >= len(sym):
    #             break
    #     if cnt == 1:
    #         # sym2[idx2 - 1] = 'V'
    #         sym2[idx2 - 1] = sym[idx2 - 1]
    #     else:
    #         sym2[idx2 - cnt:idx2] = 'N'
    #     idx2 += 1


    minlocal1 = []
    minlocal2 = []
    QS2 = []
    QS5 = []
    QS8 = []
    diff_f2 = [0]
    diff_a2 = []
    diff_f5 = [0]
    diff_a5 = []
    diff_f8 = [0]
    diff_a8 = []
    diff_fm = [0]
    diff_am = []
    diff_tb = []
    # print(namefile)
    for i in range(len(btss)):
        # print(i)
        if ecg[btss[i]] >= 0:
            d = 1
        else:
            d = -1

        if btss[i] < 30:
            s1= 0 + np.argmin(d * ecg[0:btss[i]])
        else:
            s1 = btss[i] - 30 + np.argmin(d * ecg[btss[i] - 30:btss[i]])

        if btss[i] > len(ecg) - 30:
            s2 = btss[i] + np.argmin(d * ecg[btss[i]:])
        else:
           s2 = btss[i] + np.argmin(d * ecg[btss[i]:btss[i] + 30])
        if ecg[s1]*ecg[s2]>=0:
            if abs(ecg[s1]) > abs(ecg[s2]):
                s12 = s1
            else:
                s12 = s2
        else:
            if ecg[s1]*ecg[btss[i]]<0:
                s12 = s1
            else:
                s12 = s2

        diff_tb.append(round(abs(ecg[btss][i]-ecg[s12]),6))

        cnt = 0
        f = True
        while (ecg[btss[i] - cnt] * ecg[btss[i]] > 0):
            cnt += 1
            if btss[i] - cnt < 0 or cnt > 30:
                minlocal1.append(btss[i] - cnt + 1)
                f = False
                break
        if f:
            minlocal1.append(btss[i] - cnt)
        cnt = 0
        f = True
        while (ecg[btss[i] + cnt] * ecg[btss[i]] > 0):
            cnt += 1
            if btss[i] + cnt > len(ecg)-1 or cnt > 30:
                minlocal2.append(btss[i] + cnt - 1)
                f = False
                break
        if f:
            minlocal2.append(btss[i] + cnt)
        # print(i)
        sRS = 0.8 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.8 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS2.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.5 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.5 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS5.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))

        sRS = 0.2 * (ecg[btss[i]] - ecg[minlocal2[i]]) + ecg[minlocal2[i]]
        sQR = 0.2 * (ecg[btss[i]] - ecg[minlocal1[i]]) + ecg[minlocal1[i]]
        QS8.append(np.argmin(abs(ecg[btss[i]:minlocal2[i] + 1] - sRS)) + btss[i] - (
                np.argmin(abs(ecg[minlocal1[i]:btss[i] + 1] - sQR)) + minlocal1[i]))
        if i>0 and i < len(btss) - 1:
            diff_f2.append(round(abs((QS2[i]-QS2[i-1])/QS2[i-1]),6))
            diff_a2.append(round(abs((QS2[i-1]-QS2[i])/QS2[i]),6))
            diff_f5.append(round(abs((QS5[i] - QS5[i - 1])/QS5[i - 1]),6))
            diff_a5.append(round(abs((QS5[i - 1] - QS5[i])/QS5[i]),6))
            diff_f8.append(round(abs((QS8[i] - QS8[i - 1])/QS8[i - 1]),6))
            diff_a8.append(round(abs((QS8[i - 1] - QS8[i])/QS8[i]),6))
            diff_fm.append(round((diff_f2[i]+diff_f5[i]+diff_f8[i])/3,6))
            diff_am.append(round((diff_a2[i-1]+diff_a5[i-1]+diff_a8[i-1])/3,6))
        if i==len(btss)-1:
            diff_a2.append(0)
            diff_a5.append(0)
            diff_a8.append(0)
            diff_am.append(0)



    s=0
    # diff_f = []
    # diff_a = []
    #
    # diff_f.append(np.mean(np.zeros(9)))
    # diff_a.append(round(np.mean(np.abs(np.array(
    #     [(QSv[1] / QSv[2]) ** 2, (QSs[1] / QSs[2]) ** 2, (QRs[1] / QRs[2]) ** 2, (RSs[1] / RSs[2]) ** 2,
    #      (QS2[1] / QS2[2]) ** 2, (QS5[1] / QS5[2]) ** 2,
    #      (QS8[1] / QS8[2]) ** 2]))), 5))
    #
    #
    # for i in range(1, len(QSs) - 1):
    #     diff_f.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i - 1] / QSv[i]) ** 2, (QSs[i - 1] / QSs[i]) ** 2, (QRs[i - 1] / QRs[i]) ** 2,
    #          (RSs[i - 1] / RSs[i]) ** 2, (QS2[i - 1] / QS2[i]) ** 2,
    #          (QS5[i - 1] / QS5[i]) ** 2, (QS8[i - 1] / QS8[i]) ** 2]))), 5))
    #     diff_a.append(round(np.mean(np.abs(np.array(
    #         [(QSv[i] / QSv[i + 1]) ** 2, (QSs[i] / QSs[i + 1]) ** 2, (QRs[i] / QRs[i + 1]) ** 2,
    #          (RSs[i] / RSs[i + 1]) ** 2, (QS2[i] / QS2[i + 1]) ** 2,
    #          (QS5[i] / QS5[i + 1]) ** 2, (QS8[i] / QS8[i + 1]) ** 2]))), 5))
    # diff_f.append(round(np.mean(np.abs(np.array(
    #     [(QSv[-2] / QSv[-1]) ** 2, (QSs[-2] / QSs[-1]) ** 2, (QRs[-2] / QRs[-1]) ** 2, (RSs[-2] / RSs[-1]) ** 2,
    #      (QS2[-2] / QS2[-1]) ** 2,
    #      (QS5[-2] / QS5[-1]) ** 2, (QS8[-2] / QS8[-1]) ** 2]))), 5))
    # diff_a.append(np.mean(np.zeros(9)))
    # err = []
    # err.extend(np.where(detectsym2 == 'V')[0][np.where(
    #     (sym2[np.where(detectsym2 == 'V')[0]] == 'N') | (sym2[np.where(detectsym2 == 'V')[0]] == 'O') | (
    #             sym2[np.where(detectsym2 == 'V')[0]] == 'F'))[0]] + 2)
    # err.extend(np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0][np.where(
    #     (sym2[np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0]] == 'S') | (
    #         (sym2[np.where((detectsym2 == 'N') | (detectsym2 == 'O') | (detectsym2 == 'F'))[0]] == 'V')))[0]] + 2)
    # expt= []
    # expt.extend(np.where(np.array(diff_fm)<0.15)[0] + 2)
    # expt2 = []
    # expt2.extend(np.where(np.array(diff_am) < 0.15)[0] + 2)
    err2 = []
    # return sym, bts, np.round(time,6), np.round(divba,6),np.round(ecg[bts],6),diff_tb, QS2, QS5, QS8, diff_f2, diff_a2,diff_f5, diff_a5, diff_f8, diff_a8, diff_fm, diff_am, fill_colorV, fill_colorS, nondata, err, expt, expt2, err2

    # for i in range(len(bts)):
    #     if detectsym2[i]=='V':
    #         if diff_a[i] <= 1.5:
    #         #     btss = 'N'
    #         # else:
    #             detectsym2[i] = 'S'
    # s = []
    for i in range(2,len(detectsym2)-2):
        if detectsym2[i]=='N' and divba[i] < 0.85:
            if (ecg[btss[i]] * ecg[btss[i - 1]] < 0 and ecg[btss[i]] * ecg[btss[i + 1]] < 0) or (diff_fm[i] > 0.8 and diff_am[i] > 0.8):
                detectsym2[i]='V'
                # s.append(i)

    for i in range(2, len(detectsym2) - 2):
        if detectsym2[i] == 'V':
            if (diff_fm[i] >= 0.12 and diff_am[i] >= 0.12) or max(diff_fm[i], diff_am[i]) >= 0.5 or (
                    ecg[btss[i]] * ecg[btss[i - 1]] < 0 and ecg[btss[i]] * ecg[btss[i + 1]] < 0):
                detectsym2[i] = 'V'
            else:
                detectsym2[i] = 'S'

    wfdb.wrann(record_name=str(namefile),
               extension='atr',
               sample=np.asarray(bts),
               symbol=np.asarray(sym2),
               fs=360,
               write_dir='PTK')

    # wfdb.wrann(record_name=str(namefile),
    #            extension='txt',
    #            sample=np.asarray(btss),
    #            symbol=np.asarray(detectsym2),
    #            fs=360,
    #            write_dir='PTK')

    wfdb.wrann(record_name=str(namefile),
               extension='txt',
               sample=np.asarray(btss),
               symbol=np.asarray(detectsym2),
               fs=360,
               write_dir='PTK')

DATA_DIR_FOLDER = "C:/Users/AI-DEV/Downloads/mit-bih-arrhythmia-database-1.0.0/"
NOISE_DIR_FOLDER = "C:/Users/AI-DEV/Downloads/nstdb/nstdb/"
TRAIN_DATA_DIR_STR = [101, 106, 108, 109, 112, 114, 115, 116, 118, 119, 122, 124, 201, 203, 205, 207, 208, 209, 215,
                      220, 223, 230]
VALID_DATA_DIR_STR = [100, 103, 105, 111, 113, 117, 121, 123, 200, 202, 210, 212, 213, 214, 219, 221, 222, 228, 231,
                      232, 233, 234]

# if __name__ == "__main__":
#     a, b, c, d, e, f, g, h, err, expt = main(DATA_DIR_FOLDER, 101)
if __name__ == "__main__":
    origin = True
    WriteExcel = False
    PTK = True
    if WriteExcel == False:
        # main(DATA_DIR_FOLDER, 116)
        if origin:
            if PTK:
                for i in TRAIN_DATA_DIR_STR:
                    checkPTK(DATA_DIR_FOLDER, i)
                for i in VALID_DATA_DIR_STR:
                    checkPTK(DATA_DIR_FOLDER, i)
                # ann_ptk(DATA_DIR_FOLDER, 231)
                # for i in TRAIN_DATA_DIR_STR:
                #     ann_ptk_origin(DATA_DIR_FOLDER, i)
                # for i in VALID_DATA_DIR_STR:
                #     ann_ptk_origin(DATA_DIR_FOLDER, i)
            else:
                for i in TRAIN_DATA_DIR_STR:
                    ann_origin(DATA_DIR_FOLDER, i)
                for i in VALID_DATA_DIR_STR:
                    ann_origin(DATA_DIR_FOLDER, i)
        else:
            if PTK:
                # ann_ptk(DATA_DIR_FOLDER, 116)
                for i in TRAIN_DATA_DIR_STR:
                    ann_ptk(DATA_DIR_FOLDER, i)
                for i in VALID_DATA_DIR_STR:
                    ann_ptk(DATA_DIR_FOLDER, i)
            else:
                # ann(DATA_DIR_FOLDER, 207)
                # ann(DATA_DIR_FOLDER, 232)
                # ann(DATA_DIR_FOLDER, 101)
                for i in TRAIN_DATA_DIR_STR:
                    ann(DATA_DIR_FOLDER, i)
                for i in VALID_DATA_DIR_STR:
                    ann(DATA_DIR_FOLDER, i)
        # main(DATA_DIR_FOLDER, 221)
        # a, b, c, d, e, f, g, h, err, expt = main(DATA_DIR_FOLDER, i)
    else:
        pa_sheet1_csv = []
        name_sheet1_csv = []

        fcV_sheet1 = []
        fcS_sheet1 = []
        dV_sheet1 = []
        dS_sheet1 = []

        errs_sheet1 = []
        errs2_sheet1 = []
        expts_sheet1 = []
        expts2_sheet1 = []

        pa_sheet3_csv = []
        name_sheet3_csv = []

        # pa_csv = []
        # name_csv = []
        # pa1_csv = []
        # name1_csv = []
        # name2_csv = []

        fcV_sheet3 = []
        fcS_sheet3 = []
        errs_sheet3 = []
        errs2_sheet3 = []
        expts_sheet3 = []
        expts2_sheet3 = []
        dV_sheet3 = []
        dS_sheet3 = []

        for i in TRAIN_DATA_DIR_STR:
            # a, b, c, d, e, f, g, h, err, expt, cl = main(DATA_DIR_FOLDER, i)
            sym, bts, time, divba, rr, btsv, diff_tb, QS2, QS5, QS8, diff_f2, diff_a2, diff_f5, diff_a5, diff_f8, diff_a8, diff_fm, diff_am, fill_colorV, fill_colorS, nondata, err, expt, expt2, err2, d_V, d_S = main(
                DATA_DIR_FOLDER, i)
            pa_sheet1_csv.append(sym)
            pa_sheet1_csv.append(bts)
            pa_sheet1_csv.append(time)
            pa_sheet1_csv.append(divba)
            pa_sheet1_csv.append(rr)
            pa_sheet1_csv.append(btsv)
            pa_sheet1_csv.append(diff_tb)
            pa_sheet1_csv.append(QS2)
            pa_sheet1_csv.append(QS5)
            pa_sheet1_csv.append(QS8)
            pa_sheet1_csv.append(diff_f2)
            pa_sheet1_csv.append(diff_a2)
            pa_sheet1_csv.append(diff_f5)
            pa_sheet1_csv.append(diff_a5)
            pa_sheet1_csv.append(diff_f8)
            pa_sheet1_csv.append(diff_a8)
            pa_sheet1_csv.append(diff_fm)
            pa_sheet1_csv.append(diff_am)
            pa_sheet1_csv.append(nondata)
            pa_sheet1_csv.append(nondata)

            name_sheet1_csv.append(i)
            name_sheet1_csv.append('Sample')
            name_sheet1_csv.append('Time(s)')
            name_sheet1_csv.append('RRts')
            name_sheet1_csv.append('rr')
            name_sheet1_csv.append('btsv')
            name_sheet1_csv.append('diff_tb')
            name_sheet1_csv.append('QS2')
            name_sheet1_csv.append('QS5')
            name_sheet1_csv.append('QS8')
            name_sheet1_csv.append('diff_f2')
            name_sheet1_csv.append('diff_a2')
            name_sheet1_csv.append('diff_f5')
            name_sheet1_csv.append('diff_a5')
            name_sheet1_csv.append('diff_f8')
            name_sheet1_csv.append('diff_a8')
            name_sheet1_csv.append('diff_fm')
            name_sheet1_csv.append('diff_am')
            name_sheet1_csv.append(' ')
            name_sheet1_csv.append(' ')
            fcV_sheet1.append(fill_colorV)
            fcS_sheet1.append(fill_colorS)
            dS_sheet1.append(d_S)
            dV_sheet1.append(d_V)
            errs_sheet1.append(err)
            errs2_sheet1.append(err2)
            expts_sheet1.append(expt)
            expts2_sheet1.append(expt2)

        for i in VALID_DATA_DIR_STR:
            sym, bts, time, divba, rr, btsv, diff_tb, QS2, QS5, QS8, diff_f2, diff_a2, diff_f5, diff_a5, diff_f8, diff_a8, diff_fm, diff_am, fill_colorV, fill_colorS, nondata, err, expt, expt2, err2, d_V, d_S = main(
                DATA_DIR_FOLDER, i)
            pa_sheet3_csv.append(sym)
            pa_sheet3_csv.append(bts)
            pa_sheet3_csv.append(time)
            pa_sheet3_csv.append(divba)
            pa_sheet3_csv.append(rr)
            pa_sheet3_csv.append(btsv)
            pa_sheet3_csv.append(diff_tb)
            pa_sheet3_csv.append(QS2)
            pa_sheet3_csv.append(QS5)
            pa_sheet3_csv.append(QS8)
            pa_sheet3_csv.append(diff_f2)
            pa_sheet3_csv.append(diff_a2)
            pa_sheet3_csv.append(diff_f5)
            pa_sheet3_csv.append(diff_a5)
            pa_sheet3_csv.append(diff_f8)
            pa_sheet3_csv.append(diff_a8)
            pa_sheet3_csv.append(diff_fm)
            pa_sheet3_csv.append(diff_am)
            pa_sheet3_csv.append(nondata)
            pa_sheet3_csv.append(nondata)

            name_sheet3_csv.append(i)
            name_sheet3_csv.append('Sample')
            name_sheet3_csv.append('Time(s)')
            name_sheet3_csv.append('RRts')
            name_sheet3_csv.append('rr')
            name_sheet3_csv.append('btsv')
            name_sheet3_csv.append('diff_tb')
            name_sheet3_csv.append('QS2')
            name_sheet3_csv.append('QS5')
            name_sheet3_csv.append('QS8')
            name_sheet3_csv.append('diff_f2')
            name_sheet3_csv.append('diff_a2')
            name_sheet3_csv.append('diff_f5')
            name_sheet3_csv.append('diff_a5')
            name_sheet3_csv.append('diff_f8')
            name_sheet3_csv.append('diff_a8')
            name_sheet3_csv.append('diff_fm')
            name_sheet3_csv.append('diff_am')
            name_sheet3_csv.append(' ')
            name_sheet3_csv.append(' ')
            fcV_sheet3.append(fill_colorV)
            fcS_sheet3.append(fill_colorS)
            dS_sheet3.append(d_S)
            dV_sheet3.append(d_V)
            errs_sheet3.append(err)
            errs2_sheet3.append(err2)
            expts_sheet3.append(expt)
            expts2_sheet3.append(expt2)

        maxlen = 0
        for i in range(len(pa_sheet1_csv)):
            if len(pa_sheet1_csv[i]) > maxlen:
                maxlen = len(pa_sheet1_csv[i])
        for i in range(len(pa_sheet1_csv)):
            for j in range(maxlen - len(pa_sheet1_csv[i])):
                pa_sheet1_csv[i] = np.append(pa_sheet1_csv[i], ' ')

        with open('data_morphology_sheet1.csv', 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile)

            # Sử dụng hàm zip để kết hợp dữ liệu cột theo từng cột
            csvwriter.writerow(name_sheet1_csv)
            for column_data in zip(*pa_sheet1_csv):
                csvwriter.writerow(column_data)

        maxlen = 0
        for i in range(len(pa_sheet3_csv)):
            if len(pa_sheet3_csv[i]) > maxlen:
                maxlen = len(pa_sheet3_csv[i])
        for i in range(len(pa_sheet3_csv)):
            for j in range(maxlen - len(pa_sheet3_csv[i])):
                pa_sheet3_csv[i] = np.append(pa_sheet3_csv[i], ' ')

        with open('data_morphology_sheet3.csv', 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile)

            # Sử dụng hàm zip để kết hợp dữ liệu cột theo từng cột
            csvwriter.writerow(name_sheet3_csv)
            for column_data in zip(*pa_sheet3_csv):
                csvwriter.writerow(column_data)

        # Tạo tệp Excel và tô màu ô
        # Tạo tệp Excel và tô màu ô
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "MORPHOLOGY DS1"

        # Đọc dữ liệu từ file CSV và ghi vào tệp Excel
        with open('data_morphology_sheet1.csv', 'r') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row_num, row in enumerate(csv_reader, 1):
                for col_num, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)

        for i in range(len(fcV_sheet1)):
            for j in range(len(fcV_sheet1[i])):
                row_index = fcV_sheet1[i][j]
                col_index = i * 20

                # Tô màu ô tại hàng và cột đã chỉ định
                cell_to_color = sheet.cell(row=row_index, column=col_index + 1)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 2)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 3)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 4)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 5)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 6)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 7)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 8)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 9)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 10)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 11)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 12)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 13)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 14)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 15)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 16)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 17)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet.cell(row=row_index, column=col_index + 18)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for i in range(len(fcS_sheet1)):
            for j in range(len(fcS_sheet1[i])):
                row_index = fcS_sheet1[i][j]
                col_index = i * 20
                font = Font(color="FFFFFF")
                # Tô màu ô tại hàng và cột đã chỉ định
                cell_to_color = sheet.cell(row=row_index, column=col_index + 1)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 2)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 3)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 4)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 5)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 6)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 7)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 8)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 9)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 10)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 11)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 12)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 13)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 14)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 15)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 16)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 17)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet.cell(row=row_index, column=col_index + 18)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font

        for i in range(len(dS_sheet1)):
            for j in range(len(dS_sheet1[i])):
                row_index = dS_sheet1[i][j]
                col_index = i * 20

                cell_to_color = sheet.cell(row=row_index, column=col_index + 19)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        for i in range(len(dV_sheet1)):
            for j in range(len(dV_sheet1[i])):
                row_index = dV_sheet1[i][j]
                col_index = i * 20

                cell_to_color = sheet.cell(row=row_index, column=col_index + 19)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # for i in range(len(errs_sheet1)):
        #     for j in range(len(errs_sheet1[i])):
        #         row_index = errs_sheet1[i][j]
        #         col_index = i * 20
        #
        #         # Tô màu ô tại hàng và cột đã chỉ định
        #         cell_to_color = sheet.cell(row=row_index, column=col_index + 15)
        #         cell_to_color.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        #
        for i in range(len(errs2_sheet1)):
            for j in range(len(errs2_sheet1[i])):
                row_index = errs2_sheet1[i][j]
                col_index = i * 20

                # Tô màu ô tại hàng và cột đã chỉ định
                cell_to_color = sheet.cell(row=row_index, column=col_index + 4)
                cell_to_color.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
        #
        # for i in range(len(expts_sheet1)):
        #     for j in range(len(expts_sheet1[i])):
        #         # print(i,' ', j)
        #         row_index = expts_sheet1[i][j]
        #         col_index = i * 20
        #
        #         # Tô màu ô tại hàng và cột đã chỉ định
        #         cell_to_color = sheet.cell(row=row_index, column=col_index + 16)
        #         cell_to_color.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        #
        # for i in range(len(expts2_sheet1)):
        #     for j in range(len(expts2_sheet1[i])):
        #         row_index = expts2_sheet1[i][j]
        #         col_index = i * 20
        #
        #         # Tô màu ô tại hàng và cột đã chỉ định
        #         cell_to_color = sheet.cell(row=row_index, column=col_index + 17)
        #         cell_to_color.fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

        sheet3 = workbook.create_sheet(title="MORPHOLOGY DS2")
        # Đọc dữ liệu từ file CSV và ghi vào tệp Excel
        with open('data_morphology_sheet3.csv', 'r') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row_num, row in enumerate(csv_reader, 1):
                for col_num, value in enumerate(row, 1):
                    cell = sheet3.cell(row=row_num, column=col_num, value=value)

        for i in range(len(fcV_sheet3)):
            for j in range(len(fcV_sheet3[i])):
                row_index = fcV_sheet3[i][j]
                col_index = i * 20

                # Tô màu ô tại hàng và cột đã chỉ định
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 1)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 2)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 3)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 4)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 5)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 6)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 7)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 8)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 9)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 10)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 11)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 12)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 13)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 14)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 15)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 16)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 17)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 18)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for i in range(len(fcS_sheet3)):
            for j in range(len(fcS_sheet3[i])):
                row_index = fcS_sheet3[i][j]
                col_index = i * 20
                font = Font(color="FFFFFF")
                # Tô màu ô tại hàng và cột đã chỉ định
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 1)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 2)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 3)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 4)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 5)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 6)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 7)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 8)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 9)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 10)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 11)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 12)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 13)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 14)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 15)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 16)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 17)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 18)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                cell_to_color.font = font

        # for i in range(len(errs_sheet3)):
        #     for j in range(len(errs_sheet3[i])):
        #         row_index = errs_sheet3[i][j]
        #         col_index = i * 18
        #
        #         # Tô màu ô tại hàng và cột đã chỉ định
        #         cell_to_color = sheet3.cell(row=row_index, column=col_index + 15)
        #         cell_to_color.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        #
        for i in range(len(errs2_sheet3)):
            for j in range(len(errs2_sheet3[i])):
                row_index = errs2_sheet3[i][j]
                col_index = i * 20

                # Tô màu ô tại hàng và cột đã chỉ định
                cell_to_color = sheet3.cell(row=row_index, column=col_index + 4)
                cell_to_color.fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
        #
        # for i in range(len(expts_sheet3)):
        #     for j in range(len(expts_sheet3[i])):
        #         row_index = expts_sheet3[i][j]
        #         col_index = i * 18
        #
        #         # Tô màu ô tại hàng và cột đã chỉ định
        #         cell_to_color = sheet3.cell(row=row_index, column=col_index + 16)
        #         cell_to_color.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        #
        # for i in range(len(expts2_sheet3)):
        #     for j in range(len(expts2_sheet3[i])):
        #         row_index = expts2_sheet3[i][j]
        #         col_index = i * 18
        #
        #         # Tô màu ô tại hàng và cột đã chỉ định
        #         cell_to_color = sheet3.cell(row=row_index, column=col_index + 17)
        #         cell_to_color.fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

        for i in range(len(dS_sheet3)):
            for j in range(len(dS_sheet3[i])):
                row_index = dS_sheet3[i][j]
                col_index = i * 20

                cell_to_color = sheet3.cell(row=row_index, column=col_index + 19)
                cell_to_color.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        for i in range(len(dV_sheet3)):
            for j in range(len(dV_sheet3[i])):
                row_index = dV_sheet3[i][j]
                col_index = i * 20

                cell_to_color = sheet3.cell(row=row_index, column=col_index + 19)
                cell_to_color.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        workbook.save('NSV14.xlsx')